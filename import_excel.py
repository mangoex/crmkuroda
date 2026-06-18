import asyncio
import os
import openpyxl
import unicodedata
from datetime import datetime, date
from decimal import Decimal
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select

# Set up path and import project modules
import sys
sys.path.append("/Users/renatavictoriagonzalez/Documents/miguelgespino/CRMKuroda")

from app.core.database import engine, Base, SessionLocal
from app.models.usuario import Usuario
from app.models.cotizacion import Cotizacion
from app.core.security import get_password_hash

EXCEL_PATH = "/Users/renatavictoriagonzalez/Documents/miguelgespino/CRMKuroda/MK01 2026.xlsx"
DEFAULT_PASSWORD = "kuroda2026!"

def clean_currency(val) -> Decimal:
    """Safely converts string or float currency values to Decimal."""
    if val is None:
        return Decimal("0.00")
    if isinstance(val, (int, float)):
        return Decimal(f"{val:.2f}")
    
    cleaned = str(val).replace("$", "").replace(",", "").strip()
    try:
        return Decimal(cleaned)
    except Exception:
        return Decimal("0.00")

def make_email(name: str, code: str) -> str:
    """Generates a professional name-based email from the seller's name."""
    clean_name = unicodedata.normalize('NFKD', name).encode('ASCII', 'ignore').decode('utf-8').lower()
    clean_name = ''.join(c for c in clean_name if c.isalnum() or c == ' ')
    parts = [p for p in clean_name.split() if p]
    if len(parts) >= 2:
        email_user = f"{parts[0]}.{parts[1]}"
    elif len(parts) == 1:
        email_user = parts[0]
    else:
        email_user = code.lower()
    return f"{email_user}@kuroda.com"

def parse_date(val):
    """Safely parses date fields from Excel."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    try:
        # String fallback
        return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
    except Exception:
        return None

async def import_data():
    from sqlalchemy import text
    print("Iniciando creación de tablas (si no existen)...")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
        
        # Run column migrations if tables already exist
        print("Ejecutando migraciones de columnas (ALTER TABLE)...")
        await conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS codigo_vendedor VARCHAR;"))
        await conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS nombre_completo VARCHAR;"))
        
        await conn.execute(text("ALTER TABLE cotizaciones ADD COLUMN IF NOT EXISTS numero_cotizacion VARCHAR;"))
        await conn.execute(text("ALTER TABLE cotizaciones ADD COLUMN IF NOT EXISTS fecha_registro DATE;"))
        await conn.execute(text("ALTER TABLE cotizaciones ADD COLUMN IF NOT EXISTS canal VARCHAR;"))
        await conn.execute(text("ALTER TABLE cotizaciones ADD COLUMN IF NOT EXISTS numero_factura VARCHAR;"))
        await conn.execute(text("ALTER TABLE cotizaciones ADD COLUMN IF NOT EXISTS fecha_factura DATE;"))
        await conn.execute(text("ALTER TABLE cotizaciones ADD COLUMN IF NOT EXISTS venta_perdida VARCHAR;"))
        await conn.execute(text("ALTER TABLE cotizaciones ADD COLUMN IF NOT EXISTS comentarios TEXT;"))
    print("Tablas de base de datos migradas y listas.")

    # 1. Load Excel Workbook
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Archivo de Excel no encontrado en: {EXCEL_PATH}")
        return

    print("Cargando libro de Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print("Libro cargado.")

    async with SessionLocal() as db:
        # Fetch existing quote numbers to prevent duplicates (Idempotencia)
        print("Obteniendo cotizaciones ya registradas...")
        existing_quotes_res = await db.execute(select(Cotizacion.numero_cotizacion))
        existing_quote_numbers = set(existing_quotes_res.scalars().all())
        print(f"Encontradas {len(existing_quote_numbers)} cotizaciones previas en base de datos.")

        # 2. Extract and create all sellers first
        sellers_dict = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            
            headers = rows[0]
            try:
                idx_code = headers.index('Vendedor')
                idx_name = headers.index('Nombre del Vendedor')
            except ValueError:
                continue
                
            for row in rows[1:]:
                if len(row) > idx_code and row[idx_code] is not None:
                    code = str(row[idx_code]).strip()
                    name = str(row[idx_name]).strip() if len(row) > idx_name and row[idx_name] else "Sin Nombre"
                    sellers_dict[code] = name

        print(f"Vendedores encontrados en Excel: {len(sellers_dict)}")
        
        # Populate/check users in DB
        db_sellers = {}
        hashed_pass = get_password_hash(DEFAULT_PASSWORD)
        
        for code, name in sellers_dict.items():
            email = make_email(name, code)
            
            # Check if user already exists in DB
            res = await db.execute(select(Usuario).filter(Usuario.codigo_vendedor == code))
            existing_user = res.scalars().first()
            
            if not existing_user:
                # Also double check email uniqueness
                res_email = await db.execute(select(Usuario).filter(Usuario.email == email))
                if res_email.scalars().first():
                    # Append code suffix to prevent duplicates
                    email = f"{email.split('@')[0]}_{code.lower()}@kuroda.com"
                
                print(f"Registrando Vendedor: {name} ({code}) -> Email: {email}")
                new_seller = Usuario(
                    email=email,
                    hashed_password=hashed_pass,
                    rol="vendedor",
                    telefono_whatsapp=None, # will be linked manually
                    codigo_vendedor=code,
                    nombre_completo=name
                )
                db.add(new_seller)
                await db.flush() # obtain ID
                db_sellers[code] = new_seller.id
            else:
                db_sellers[code] = existing_user.id
                
        await db.commit()
        print("Vendedores inicializados en la base de datos.")

        # 3. Import Cotizaciones sheet by sheet
        total_imported = 0
        total_skipped = 0
        
        for sheet_name in wb.sheetnames:
            print(f"Procesando hoja: {sheet_name}...")
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            
            headers = rows[0]
            try:
                idx_fecha_reg = headers.index('Fecha de Registro')
                idx_num_cot = headers.index('Numero de Cotizacion')
                idx_canal = headers.index('Canal')
                idx_vendedor = headers.index('Vendedor')
                idx_num_cli = headers.index('Numero del Cliente')
                idx_nom_cli = headers.index('Nombre del Cliente')
                idx_tel = headers.index('Numero de Telefono')
                idx_cel = headers.index('Numero de Celular')
                idx_email = headers.index('Direccion Correo Electronico')
                idx_importe = headers.index('Importe Cotizado c/IVA')
                idx_num_fac = headers.index('Numero de Factura')
                idx_fecha_fac = headers.index('Fecha de Factura')
                idx_perdida = headers.index('Venta perdida?')
                idx_coment = headers.index('Comentarios')
            except ValueError as e:
                print(f"Ignorando hoja {sheet_name}: Falta columna clave ({e})")
                continue

            cotizaciones_batch = []
            for i, row in enumerate(rows[1:]):
                if len(row) <= idx_vendedor or row[idx_vendedor] is None:
                    continue
                
                vend_code = str(row[idx_vendedor]).strip()
                vendedor_id = db_sellers.get(vend_code)
                if not vendedor_id:
                    continue
                
                # Check for duplicate quote number
                num_cot = str(int(row[idx_num_cot])) if isinstance(row[idx_num_cot], (int, float)) else str(row[idx_num_cot]) if row[idx_num_cot] else None
                if num_cot and num_cot in existing_quote_numbers:
                    total_skipped += 1
                    continue
                
                # Parse quote columns
                fecha_reg = parse_date(row[idx_fecha_reg])
                canal = str(row[idx_canal]).strip() if row[idx_canal] is not None else None
                
                num_cli = str(int(row[idx_num_cli])) if isinstance(row[idx_num_cli], (int, float)) else str(row[idx_num_cli]) if row[idx_num_cli] else None
                nom_cli = str(row[idx_nom_cli]).strip() if row[idx_nom_cli] else "Cliente Desconocido"
                
                tel = str(row[idx_tel]).strip() if row[idx_tel] else None
                cel = str(row[idx_cel]).strip() if row[idx_cel] else None
                email_cli = str(row[idx_email]).strip() if row[idx_email] else None
                
                total = clean_currency(row[idx_importe])
                
                num_fac = str(int(row[idx_num_fac])) if isinstance(row[idx_num_fac], (int, float)) else str(row[idx_num_fac]) if row[idx_num_fac] else None
                fecha_fac = parse_date(row[idx_fecha_fac])
                
                venta_perdida = str(row[idx_perdida]).strip() if row[idx_perdida] else "No"
                comentarios = str(row[idx_coment]).strip() if row[idx_coment] else None

                # Build JSON metadata structures
                datos_contacto = {
                    "numero_cliente": num_cli,
                    "telefono": tel,
                    "celular": cel,
                    "email": email_cli
                }
                
                items = [
                    {
                        "producto": "Cotización consolidada",
                        "cantidad": 1,
                        "precio_unitario": float(total)
                    }
                ]
                
                texto_propuesta = comentarios or f"Cotización formal emitida el {fecha_reg} para el cliente {nom_cli} por un importe total de ${total}."

                new_cot = Cotizacion(
                    vendedor_id=vendedor_id,
                    cliente_nombre=nom_cli,
                    datos_contacto=datos_contacto,
                    items=items,
                    total=total,
                    texto_propuesta=texto_propuesta,
                    numero_cotizacion=num_cot,
                    fecha_registro=fecha_reg,
                    canal=canal,
                    numero_factura=num_fac,
                    fecha_factura=fecha_fac,
                    venta_perdida=venta_perdida,
                    comentarios=comentarios
                )
                cotizaciones_batch.append(new_cot)
                
                # Keep local set updated to prevent duplicates within the sheet
                if num_cot:
                    existing_quote_numbers.add(num_cot)
                
                if len(cotizaciones_batch) >= 500:
                    db.add_all(cotizaciones_batch)
                    await db.commit()
                    total_imported += len(cotizaciones_batch)
                    print(f"   Importadas {total_imported} cotizaciones...")
                    cotizaciones_batch = []
            
            # Insert remaining in last batch
            if cotizaciones_batch:
                db.add_all(cotizaciones_batch)
                await db.commit()
                total_imported += len(cotizaciones_batch)
                print(f"   Importadas {total_imported} cotizaciones...")

        print(f"\n¡IMPORTACIÓN FINALIZADA!")
        print(f"- Cotizaciones importadas: {total_imported}")
        print(f"- Cotizaciones duplicadas omitidas: {total_skipped}")

if __name__ == "__main__":
    asyncio.run(import_data())
