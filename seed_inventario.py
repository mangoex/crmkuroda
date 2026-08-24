import io
import os
import logging
import openpyxl
from sqlalchemy.future import select
from sqlalchemy import func, insert, delete
from app.core.database import SessionLocal, engine, Base
from app.models.inventario_abcf import InventarioAbcf
from app.api.v1.inventario_abcf import _normalize_header, _header_index, _row_value, _as_float

logger = logging.getLogger(__name__)

EXCEL_PATHS = [
    os.path.join(os.path.dirname(__file__), "Inventario MKS D.XLSX"),
    os.path.join(os.path.dirname(__file__), "..", "Inventario MKS D.XLSX"),
    os.path.join(os.path.dirname(__file__), "Inventario MKS al 29.06.26.XLSX"),
    os.path.join(os.path.dirname(__file__), "..", "Inventario MKS al 29.06.26.XLSX"),
    "Inventario MKS D.XLSX",
    "../Inventario MKS D.XLSX",
    "Inventario MKS al 29.06.26.XLSX",
    "../Inventario MKS al 29.06.26.XLSX",
]


def find_inventario_excel():
    for p in EXCEL_PATHS:
        abs_p = os.path.abspath(p)
        if os.path.exists(abs_p):
            return abs_p
    return None


def parse_inventario_rows_from_workbook(wb) -> list[dict]:
    """Extrae y parsea filas con precios y ubicaciones desde un libro openpyxl."""
    records = []
    seen_keys = set()

    for ws in wb.worksheets:
        if ws.sheet_state == "hidden":
            continue

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [_normalize_header(value) for value in header_row]
        
        indices = {
            "centro": _header_index(headers, "centro", "sucursal", "centro distribucion", "nombre centro"),
            "almacen": _header_index(headers, "almacen", "almacen origen"),
            "numero_proveedor": _header_index(headers, "numero proveedor", "codigo proveedor", "proveedor codigo", "numero de proveedor"),
            "nombre_proveedor": _header_index(headers, "nombre proveedor", "proveedor", "razon social proveedor", "nombre del proveedor"),
            "abc_f": _header_index(
                headers,
                "abc+f",
                "abcf",
                "codigo abcf",
                "clasificacion abcf",
                "indicador abc+frecuencia de venta",
                "indicador abcf frecuencia de venta",
                "d",
            ),
            "codigo_material": _header_index(headers, "codigo material", "clave material", "codigo producto", "clave producto", "sku"),
            "descripcion_material": _header_index(headers, "descripcion del material", "descripcion material", "descripcion producto", "descripcion", "producto"),
            "cantidad_propia": _header_index(headers, "cantidad propia", "cant propia", "inventario disponible", "existencia propia", "disponible"),
            "existencia_consignacion": _header_index(headers, "existencia consignacion", "inv consig", "inventario consignacion", "existencia en consignacion de proveedore", "existencia en consignacion de proveedores"),
            "entregas_pendientes": _header_index(headers, "entregas pendientes"),
            "existencia_transito": _header_index(headers, "existencia transito", "transito"),
            "existencia_bloqueada": _header_index(headers, "existencia bloqueada", "bloqueada"),
            "existencia_control_calidad": _header_index(headers, "existencia control calidad", "control calidad"),
            "umb": _header_index(headers, "umb", "unidad medida", "unidad de medida base"),
            "costo_promedio_unitario": _header_index(
                headers,
                "precio venta",
                "precio de venta",
                "precio lista",
                "precio unitario",
                "precio comercial",
                "precio",
                "pvp",
                "costo promedio unitario",
                "precio promedio",
                "costo promedio",
                "precio prom",
                "precio promocion",
            ),
            "importe_inventario_propio": _header_index(headers, "importe inventario propio", "importe inv", "importe de inventario propio"),
            "valor_consignacion_proveedor": _header_index(headers, "valor consignacion proveedor", "valor de consignacion proveedor"),
            "ubicacion": _header_index(headers, "ubicacion", "localizacion"),
            "grupo_materiales": _header_index(headers, "grupo materiales"),
            "descrip_gpo_materiales": _header_index(headers, "descripcion grupo materiales", "descrip gpo materiales"),
            "codigo_anterior_material": _header_index(headers, "codigo anterior material"),
            "abc": _header_index(headers, "indicador abc", "abc"),
            "fecha_ultimo_inventario": _header_index(headers, "fecha ultimo inventario", "fecha del ultimo inventario ciclico"),
        }

        # Ignorar hojas que no tienen código ni descripción ni centro
        if indices["codigo_material"] is None or indices["centro"] is None:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not _row_value(row, indices["centro"], 0):
                continue

            try:
                c_propia = _as_float(_row_value(row, indices["cantidad_propia"], 7), 0.0)
                e_consig = _as_float(_row_value(row, indices["existencia_consignacion"], 8), 0.0)

                if c_propia == 0.0 and e_consig == 0.0:
                    continue

                c_unitario = _as_float(_row_value(row, indices["costo_promedio_unitario"], 14))
                imp_propio = _as_float(_row_value(row, indices["importe_inventario_propio"], 15))
                if (c_unitario is None or c_unitario == 0.0) and imp_propio and imp_propio > 0 and c_propia > 0:
                    c_unitario = round(imp_propio / c_propia, 2)

                centro_val = str(_row_value(row, indices["centro"], 0)) if _row_value(row, indices["centro"], 0) is not None else None
                cod_mat = str(_row_value(row, indices["codigo_material"], 1)) if _row_value(row, indices["codigo_material"], 1) is not None else None
                almacen_val = str(_row_value(row, indices["almacen"])) if _row_value(row, indices["almacen"]) is not None else None

                # Clave única para evitar duplicados entre hojas idénticas
                dedup_key = (centro_val, almacen_val, cod_mat)
                if dedup_key in seen_keys:
                    continue
                seen_keys.add(dedup_key)

                records.append({
                    "nombre_centro": centro_val,
                    "almacen": almacen_val,
                    "numero_proveedor": str(_row_value(row, indices["numero_proveedor"])) if _row_value(row, indices["numero_proveedor"]) is not None else None,
                    "nombre_proveedor": str(_row_value(row, indices["nombre_proveedor"], 2)) if _row_value(row, indices["nombre_proveedor"], 2) is not None else None,
                    "abc_f": str(_row_value(row, indices["abc_f"])) if _row_value(row, indices["abc_f"]) is not None else None,
                    "codigo_material": cod_mat,
                    "descripcion_material": str(_row_value(row, indices["descripcion_material"], 3)) if _row_value(row, indices["descripcion_material"], 3) is not None else None,
                    "cantidad_propia": c_propia,
                    "existencia_consignacion": e_consig,
                    "entregas_pendientes": _as_float(_row_value(row, indices["entregas_pendientes"], 9)),
                    "existencia_transito": _as_float(_row_value(row, indices["existencia_transito"], 10)),
                    "existencia_bloqueada": _as_float(_row_value(row, indices["existencia_bloqueada"], 11)),
                    "existencia_control_calidad": _as_float(_row_value(row, indices["existencia_control_calidad"], 12)),
                    "umb": str(_row_value(row, indices["umb"], 13)) if _row_value(row, indices["umb"], 13) is not None else None,
                    "costo_promedio_unitario": c_unitario,
                    "importe_inventario_propio": imp_propio,
                    "valor_consignacion_proveedor": _as_float(_row_value(row, indices["valor_consignacion_proveedor"], 16)),
                    "ubicacion": str(_row_value(row, indices["ubicacion"], 17)) if _row_value(row, indices["ubicacion"], 17) is not None else None,
                    "grupo_materiales": str(_row_value(row, indices["grupo_materiales"], 18)) if _row_value(row, indices["grupo_materiales"], 18) is not None else None,
                    "descrip_gpo_materiales": str(_row_value(row, indices["descrip_gpo_materiales"], 19)) if _row_value(row, indices["descrip_gpo_materiales"], 19) is not None else None,
                    "codigo_anterior_material": str(_row_value(row, indices["codigo_anterior_material"], 20)) if _row_value(row, indices["codigo_anterior_material"], 20) is not None else None,
                    "abc": str(_row_value(row, indices["abc"], 21)) if _row_value(row, indices["abc"], 21) is not None else None,
                    "fecha_ultimo_inventario": str(_row_value(row, indices["fecha_ultimo_inventario"], 22)) if _row_value(row, indices["fecha_ultimo_inventario"], 22) is not None else None,
                })
            except Exception as e:
                continue

    return records


async def seed_inventario_from_excel(force=False):
    """Siembra o repara automáticamente el inventario ABC+F con precios y ubicaciones válidas."""
    excel_file = find_inventario_excel()
    if not excel_file:
        logger.info("No se encontró archivo Excel de Inventario para siembra automática.")
        return 0

    async with SessionLocal() as session:
        if not force:
            count_res = await session.execute(select(func.count(InventarioAbcf.id)))
            existing_count = count_res.scalar() or 0

            # Verificar si los registros existentes carecen de precios
            if existing_count > 0:
                valid_price_res = await session.execute(
                    select(func.count(InventarioAbcf.id)).where(
                        InventarioAbcf.costo_promedio_unitario.isnot(None),
                        InventarioAbcf.costo_promedio_unitario > 0
                    )
                )
                valid_price_count = valid_price_res.scalar() or 0
                
                # Si ya tiene registros con precio, no es necesario resembrar
                if valid_price_count > 0:
                    logger.info(f"Inventario ABC+F ya contiene {existing_count} registros ({valid_price_count} con precio válido). Se omite siembra.")
                    return existing_count
                
                logger.info(f"Inventario existente contiene 0 precios válidos. Re-sembrando automáticamente...")

        print(f"Cargando Inventario ABC+F desde: {excel_file}")
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        records = parse_inventario_rows_from_workbook(wb)

        if not records:
            return 0

        # Eliminar registros anteriores si se está forzando o reparando
        await session.execute(delete(InventarioAbcf))
        await session.commit()

        batch_size = 1000
        total_inserted = 0
        for i in range(0, len(records), batch_size):
            batch = records[i:i + batch_size]
            await session.execute(insert(InventarioAbcf), batch)
            await session.commit()
            total_inserted += len(batch)

        print(f"¡Siembra de Inventario ABC+F completada! ({total_inserted} registros con precios y ubicaciones)")
        return total_inserted
