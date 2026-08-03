from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import delete
import io
import openpyxl
from typing import Optional

from app.core.database import get_db
from app.core.security import RoleChecker, get_current_user
from app.models.inventario_abcf import InventarioAbcf
from app.models.usuario import Usuario
from app.services.actualizaciones_datos import registrar_actualizacion_datos

router = APIRouter()

require_admin = RoleChecker(["admin", "gerente"])


def _normalize_header(value) -> str:
    """Normaliza encabezados de Excel para tolerar acentos y cambios menores."""
    import unicodedata

    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return "".join(char for char in text.lower() if char.isalnum())


def _header_index(headers: list[str], *aliases: str) -> Optional[int]:
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    return next((index for index, header in enumerate(headers) if header in normalized_aliases), None)


def _row_value(row, index: Optional[int], fallback: Optional[int] = None):
    target = index if index is not None else fallback
    return row[target] if target is not None and len(row) > target else None


def _as_float(value, default=None):
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    val_str = str(value).replace("$", "").replace(",", "").strip()
    try:
        return float(val_str)
    except ValueError:
        return default

@router.get("/")
async def list_inventario(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(InventarioAbcf))
    inventarios = result.scalars().all()
    return {"status": "success", "data": [i.to_dict() for i in inventarios]}

@router.post("/upload", status_code=status.HTTP_201_CREATED)
async def upload_inventario(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(require_admin)
):
    if not file.filename.endswith(".xlsx") and not file.filename.endswith(".XLSX"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")
    
    contents = await file.read()
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        # Eliminar inventario anterior
        await db.execute(delete(InventarioAbcf))
        
        rows_added = 0
        
        for ws in wb.worksheets:
            if ws.sheet_state == 'hidden':
                continue

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = [_normalize_header(value) for value in header_row]
            indices = {
                "centro": _header_index(headers, "centro", "sucursal", "centro distribucion", "nombre centro"),
                "almacen": _header_index(headers, "almacen", "almacen origen"),
                "numero_proveedor": _header_index(headers, "numero proveedor", "codigo proveedor", "proveedor codigo", "numero de proveedor"),
                "nombre_proveedor": _header_index(headers, "nombre proveedor", "proveedor", "razon social proveedor", "nombre del proveedor"),
                # D representa el indicador ABC+Frecuencia de Venta, no la clave de material.
                "abc_f": _header_index(
                    headers,
                    "abc+f",
                    "abcf",
                    "codigo abcf",
                    "clasificacion abcf",
                    "indicador abc+frecuencia de venta",
                    "indicador abcf frecuencia de venta",
                    "d",
                ),
                "codigo_material": _header_index(headers, "codigo material", "clave material", "codigo producto", "clave producto", "sku"),
                "descripcion_material": _header_index(headers, "descripcion material", "descripcion producto", "descripcion", "producto"),
                "cantidad_propia": _header_index(headers, "cantidad propia", "cant propia", "inventario disponible", "existencia propia", "disponible"),
                "existencia_consignacion": _header_index(headers, "existencia consignacion", "inv consig", "inventario consignacion", "existencia en consignacion de proveedore", "existencia en consignacion de proveedores"),
                "entregas_pendientes": _header_index(headers, "entregas pendientes"),
                "existencia_transito": _header_index(headers, "existencia transito", "transito"),
                "existencia_bloqueada": _header_index(headers, "existencia bloqueada", "bloqueada"),
                "existencia_control_calidad": _header_index(headers, "existencia control calidad", "control calidad"),
                "umb": _header_index(headers, "umb", "unidad medida"),
                "costo_promedio_unitario": _header_index(headers, "costo promedio unitario", "precio promedio", "costo promedio"),
                "importe_inventario_propio": _header_index(headers, "importe inventario propio", "importe inv", "importe de inventario propio"),
                "valor_consignacion_proveedor": _header_index(headers, "valor consignacion proveedor"),
                "ubicacion": _header_index(headers, "ubicacion", "localizacion"),
                "grupo_materiales": _header_index(headers, "grupo materiales"),
                "descrip_gpo_materiales": _header_index(headers, "descripcion grupo materiales", "descrip gpo materiales"),
                "codigo_anterior_material": _header_index(headers, "codigo anterior material"),
                "abc": _header_index(headers, "abc"),
                "fecha_ultimo_inventario": _header_index(headers, "fecha ultimo inventario", "fecha del ultimo inventario ciclico"),
            }

            iter_rows = ws.iter_rows(min_row=2, values_only=True)
            for row in iter_rows:
                if not row or not _row_value(row, indices["centro"], 0):
                    continue

                try:
                    c_propia = _as_float(_row_value(row, indices["cantidad_propia"], 7), 0.0)
                    e_consig = _as_float(_row_value(row, indices["existencia_consignacion"], 8), 0.0)
                    
                    if c_propia == 0.0 and e_consig == 0.0:
                        continue
                        
                    inv = InventarioAbcf(
                        nombre_centro=str(_row_value(row, indices["centro"], 0)) if _row_value(row, indices["centro"], 0) is not None else None,
                        almacen=str(_row_value(row, indices["almacen"])) if _row_value(row, indices["almacen"]) is not None else None,
                        numero_proveedor=str(_row_value(row, indices["numero_proveedor"])) if _row_value(row, indices["numero_proveedor"]) is not None else None,
                        nombre_proveedor=str(_row_value(row, indices["nombre_proveedor"], 2)) if _row_value(row, indices["nombre_proveedor"], 2) is not None else None,
                        abc_f=str(_row_value(row, indices["abc_f"])) if _row_value(row, indices["abc_f"]) is not None else None,
                        codigo_material=str(_row_value(row, indices["codigo_material"], 1)) if _row_value(row, indices["codigo_material"], 1) is not None else None,
                        descripcion_material=str(_row_value(row, indices["descripcion_material"], 3)) if _row_value(row, indices["descripcion_material"], 3) is not None else None,
                        cantidad_propia=c_propia,
                        existencia_consignacion=e_consig,
                        entregas_pendientes=_as_float(_row_value(row, indices["entregas_pendientes"], 9)),
                        existencia_transito=_as_float(_row_value(row, indices["existencia_transito"], 10)),
                        existencia_bloqueada=_as_float(_row_value(row, indices["existencia_bloqueada"], 11)),
                        existencia_control_calidad=_as_float(_row_value(row, indices["existencia_control_calidad"], 12)),
                        umb=str(_row_value(row, indices["umb"], 13)) if _row_value(row, indices["umb"], 13) is not None else None,
                        costo_promedio_unitario=_as_float(_row_value(row, indices["costo_promedio_unitario"], 14)),
                        importe_inventario_propio=_as_float(_row_value(row, indices["importe_inventario_propio"], 15)),
                        valor_consignacion_proveedor=_as_float(_row_value(row, indices["valor_consignacion_proveedor"], 16)),
                        ubicacion=str(_row_value(row, indices["ubicacion"], 17)) if _row_value(row, indices["ubicacion"], 17) is not None else None,
                        grupo_materiales=str(_row_value(row, indices["grupo_materiales"], 18)) if _row_value(row, indices["grupo_materiales"], 18) is not None else None,
                        descrip_gpo_materiales=str(_row_value(row, indices["descrip_gpo_materiales"], 19)) if _row_value(row, indices["descrip_gpo_materiales"], 19) is not None else None,
                        codigo_anterior_material=str(_row_value(row, indices["codigo_anterior_material"], 20)) if _row_value(row, indices["codigo_anterior_material"], 20) is not None else None,
                        abc=str(_row_value(row, indices["abc"], 21)) if _row_value(row, indices["abc"], 21) is not None else None,
                        fecha_ultimo_inventario=str(_row_value(row, indices["fecha_ultimo_inventario"], 22)) if _row_value(row, indices["fecha_ultimo_inventario"], 22) is not None else None
                    )
                    db.add(inv)
                    rows_added += 1
                except Exception as row_error:
                    print(f"Error parseando fila: {row_error}")
                    continue
            
        await registrar_actualizacion_datos(db, "inventario-abcf", current_user.id)
        await db.commit()
        return {"status": "success", "message": f"Se han cargado {rows_added} registros de inventario exitosamente."}
        
    except Exception as e:
        await db.rollback()
        print(f"Error general procesando archivo de inventario: {e}")
        raise HTTPException(status_code=500, detail=f"Error procesando el archivo: {str(e)}")
