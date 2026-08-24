import csv
from datetime import date, datetime
import io
import openpyxl
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import delete

from app.core.database import get_db
from app.core.security import RoleChecker, get_current_user
from app.models.promocion import Promocion
from app.models.usuario import Usuario
from app.models.cotizacion import Cotizacion
from app.models.cotizacion_detalle import CotizacionItem
from app.services.commercial_analytics import find_clients_for_promotion, safe_phone_href
from app.services.jerarquia import get_ids_vendedores_visibles
from app.services.actualizaciones_datos import registrar_actualizacion_datos

router = APIRouter()

require_admin = RoleChecker(["admin", "gerente", "marketing"])


def _normalize_header(value) -> str:
    """Normaliza encabezados para tolerar acentos, mayusculas y caracteres especiales."""
    import unicodedata

    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return "".join(char for char in text.lower() if char.isalnum())


def _header_index(headers: list[str], *aliases: str) -> Optional[int]:
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    return next((index for index, header in enumerate(headers) if header in normalized_aliases), None)


def _row_value(row, index: Optional[int], fallback: Optional[int] = None):
    target = index if index is not None else fallback
    return row[target] if target is not None and len(row) > target else None


def _as_float(value, default=None):
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    val_str = str(value).replace("$", "").replace(",", "").strip()
    try:
        return float(val_str)
    except ValueError:
        return default


def _as_datetime(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        val_str = value.strip()
        for fmt in (
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y/%m/%d",
            "%Y-%m-%d %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%d.%m.%Y",
            "%d.%m.%Y %H:%M:%S",
        ):
            try:
                return datetime.strptime(val_str, fmt)
            except ValueError:
                continue
    return None


@router.get("/")
async def list_promociones(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Promocion))
    promociones = result.scalars().all()
    return {"status": "success", "data": [p.to_dict() for p in promociones]}


@router.post("/upload", status_code=status.HTTP_201_CREATED)
async def upload_promociones(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(require_admin)
):
    filename_lower = (file.filename or "").lower()
    if not (filename_lower.endswith(".xlsx") or filename_lower.endswith(".csv")):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx) o CSV (.csv)")
    
    contents = await file.read()
    
    try:
        rows_to_process = []
        indices = {}

        if filename_lower.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            
            # Buscar la hoja que contenga las columnas de promociones
            best_sheet = None
            best_score = -1
            best_indices = None

            for ws in wb.worksheets:
                if ws.sheet_state == "hidden":
                    continue
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                headers = [_normalize_header(value) for value in header_row]
                
                sheet_indices = {
                    "centro": _header_index(headers, "centro", "sucursal", "centro distribucion", "nombre centro"),
                    "descrip_gpo_materiales": _header_index(headers, "descrip gpo materiales", "descripcion grupo materiales", "grupo materiales", "descripcion de grupo materiales", "linea"),
                    "indicador_abc": _header_index(headers, "indicador abc+frecuencia de venta", "indicador abcf frecuencia de venta", "indicador abc", "abc+f", "abcf", "abc"),
                    "codigo_material": _header_index(headers, "codigo material", "clave material", "codigo producto", "clave producto", "sku", "material"),
                    "descripcion_material": _header_index(headers, "descripcion del material", "descripcion material", "descripcion producto", "descripcion", "producto"),
                    "unidad_medida": _header_index(headers, "unidad de medida base", "unidad medida", "unidad de medida", "umb", "unidad"),
                    "costo_promedio": _header_index(headers, "costo promedio unitario", "costo promedio", "costo prom unitario", "costo prom"),
                    "costo_promedio_moneda": _header_index(headers, "costo promedio unitario moneda de venta", "costo promedio moneda", "costo promedio moneda de venta"),
                    "costo_estandar": _header_index(headers, "costo estandar", "costo std"),
                    "precio_promocion": _header_index(headers, "precio efectivo promocion", "precio promocion", "precio promo", "precio prom", "precio de promocion", "precio especial"),
                    "moneda": _header_index(headers, "moneda", "divisa"),
                    "valido_hasta": _header_index(headers, "valido hasta promocion", "valido hasta", "fecha fin", "vigencia hasta", "vigencia"),
                    "costo_estandar_promocion": _header_index(headers, "costo estandar promocion", "costo estandar promo", "costo std promo"),
                    "margen_promocion": _header_index(headers, "margen promocion", "margen promo", "margen"),
                    "proveedor": _header_index(headers, "proveedor", "nombre del proveedor", "nombre proveedor", "razon social proveedor"),
                    "inventario_disponible": _header_index(headers, "inventario disponible", "existencia propia", "cantidad propia", "disponible", "existencia"),
                }

                score = sum(1 for v in sheet_indices.values() if v is not None)
                if sheet_indices["precio_promocion"] is not None:
                    score += 10
                if sheet_indices["costo_estandar_promocion"] is not None or sheet_indices["margen_promocion"] is not None:
                    score += 5
                if sheet_indices["descrip_gpo_materiales"] is not None:
                    score += 3

                if score > best_score:
                    best_score = score
                    best_sheet = ws
                    best_indices = sheet_indices

            if not best_sheet:
                best_sheet = wb.active
                best_indices = {}

            ws = best_sheet
            indices = best_indices or {}
            rows_to_process = list(ws.iter_rows(min_row=2, values_only=True))

        else:  # CSV
            text_data = contents.decode("utf-8-sig", errors="replace")
            sample = text_data[:2048]
            delimiter = ";" if sample.count(";") > sample.count(",") else ","
            reader = csv.reader(io.StringIO(text_data), delimiter=delimiter)
            
            header_row = next(reader, None)
            if header_row:
                headers = [_normalize_header(value) for value in header_row]
                indices = {
                    "centro": _header_index(headers, "centro", "sucursal", "centro distribucion", "nombre centro"),
                    "descrip_gpo_materiales": _header_index(headers, "descrip gpo materiales", "descripcion grupo materiales", "grupo materiales", "descripcion de grupo materiales", "linea"),
                    "indicador_abc": _header_index(headers, "indicador abc+frecuencia de venta", "indicador abcf frecuencia de venta", "indicador abc", "abc+f", "abcf", "abc"),
                    "codigo_material": _header_index(headers, "codigo material", "clave material", "codigo producto", "clave producto", "sku", "material"),
                    "descripcion_material": _header_index(headers, "descripcion del material", "descripcion material", "descripcion producto", "descripcion", "producto"),
                    "unidad_medida": _header_index(headers, "unidad de medida base", "unidad medida", "unidad de medida", "umb", "unidad"),
                    "costo_promedio": _header_index(headers, "costo promedio unitario", "costo promedio", "costo prom unitario", "costo prom"),
                    "costo_promedio_moneda": _header_index(headers, "costo promedio unitario moneda de venta", "costo promedio moneda", "costo promedio moneda de venta"),
                    "costo_estandar": _header_index(headers, "costo estandar", "costo std"),
                    "precio_promocion": _header_index(headers, "precio efectivo promocion", "precio promocion", "precio promo", "precio prom", "precio de promocion", "precio especial"),
                    "moneda": _header_index(headers, "moneda", "divisa"),
                    "valido_hasta": _header_index(headers, "valido hasta promocion", "valido hasta", "fecha fin", "vigencia hasta", "vigencia"),
                    "costo_estandar_promocion": _header_index(headers, "costo estandar promocion", "costo estandar promo", "costo std promo"),
                    "margen_promocion": _header_index(headers, "margen promocion", "margen promo", "margen"),
                    "proveedor": _header_index(headers, "proveedor", "nombre del proveedor", "nombre proveedor", "razon social proveedor"),
                    "inventario_disponible": _header_index(headers, "inventario disponible", "existencia propia", "cantidad propia", "disponible", "existencia"),
                }
                rows_to_process = list(reader)

        # Eliminar promociones anteriores
        await db.execute(delete(Promocion))
        
        rows_added = 0
        batch = []
        BATCH_SIZE = 1000

        for row in rows_to_process:
            if not row:
                continue
            
            centro_val = _row_value(row, indices.get("centro"), 0)
            cod_mat_val = _row_value(row, indices.get("codigo_material"), 3)
            
            if centro_val is None and cod_mat_val is None:
                continue
            
            if str(centro_val).strip().lower() == "centro" or str(cod_mat_val).strip().lower() in ("codigo material", "material"):
                continue

            promocion = Promocion(
                centro=str(centro_val).strip() if centro_val is not None else None,
                descrip_gpo_materiales=str(_row_value(row, indices.get("descrip_gpo_materiales"), 1)).strip() if _row_value(row, indices.get("descrip_gpo_materiales"), 1) is not None else None,
                indicador_abc=str(_row_value(row, indices.get("indicador_abc"), 2)).strip() if _row_value(row, indices.get("indicador_abc"), 2) is not None else None,
                codigo_material=str(cod_mat_val).strip() if cod_mat_val is not None else None,
                descripcion_material=str(_row_value(row, indices.get("descripcion_material"), 4)).strip() if _row_value(row, indices.get("descripcion_material"), 4) is not None else None,
                unidad_medida=str(_row_value(row, indices.get("unidad_medida"), 5)).strip() if _row_value(row, indices.get("unidad_medida"), 5) is not None else None,
                costo_promedio=_as_float(_row_value(row, indices.get("costo_promedio"), 6)),
                costo_promedio_moneda=_as_float(_row_value(row, indices.get("costo_promedio_moneda"), 7)),
                costo_estandar=_as_float(_row_value(row, indices.get("costo_estandar"), 8)),
                precio_promocion=_as_float(_row_value(row, indices.get("precio_promocion"), 9)),
                moneda=str(_row_value(row, indices.get("moneda"), 10)).strip() if _row_value(row, indices.get("moneda"), 10) is not None else None,
                valido_hasta=_as_datetime(_row_value(row, indices.get("valido_hasta"), 11)),
                costo_estandar_promocion=_as_float(_row_value(row, indices.get("costo_estandar_promocion"), 12)),
                margen_promocion=_as_float(_row_value(row, indices.get("margen_promocion"), 13)),
                proveedor=str(_row_value(row, indices.get("proveedor"), 14)).strip() if _row_value(row, indices.get("proveedor"), 14) is not None else None,
                inventario_disponible=_as_float(_row_value(row, indices.get("inventario_disponible"), 15)),
            )
            batch.append(promocion)
            rows_added += 1

            if len(batch) >= BATCH_SIZE:
                db.add_all(batch)
                batch = []

        if batch:
            db.add_all(batch)

        await registrar_actualizacion_datos(db, "promociones", current_user.id)
        await db.commit()
        return {"status": "success", "message": f"Se han cargado {rows_added} promociones exitosamente."}
        
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error procesando el archivo: {str(e)}")


@router.get("/{promocion_id}/clientes-potenciales", status_code=status.HTTP_200_OK)
async def get_promotion_potential_clients(
    promocion_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """
    Find clients who previously purchased the material that is now on promotion.

    Returns a list of clients with their contact info and purchase history.
    Salespeople only see clients linked to their quotes.
    """
    if current_user.rol == "soporte":
        raise HTTPException(
            status_code=403,
            detail="El rol de soporte no tiene acceso a esta funcionalidad.",
        )

    # Fetch the promotion.
    promotion = (
        await db.execute(select(Promocion).where(Promocion.id == promocion_id))
    ).scalars().first()
    if not promotion:
        raise HTTPException(status_code=404, detail="La promoción no existe.")

    # Fetch items + their parent quotes via join.
    from sqlalchemy import and_
    query = (
        select(CotizacionItem, Cotizacion)
        .join(Cotizacion, CotizacionItem.cotizacion_id == Cotizacion.id)
    )

    # Apply seller visibility filter for vendedor role.
    if current_user.rol == "vendedor":
        from sqlalchemy import or_, func
        ids_visibles = await get_ids_vendedores_visibles(db, current_user)
        if ids_visibles is not None:
            visible_users = (
                await db.execute(select(Usuario).where(Usuario.id.in_(ids_visibles)))
            ).scalars().all()
            visible_names = [
                user.nombre_completo.strip().upper()
                for user in visible_users
                if user.nombre_completo and user.nombre_completo.strip()
            ]
            seller_filters = [Cotizacion.vendedor_id.in_(ids_visibles)]
            if visible_names:
                seller_filters.append(
                    and_(
                        Cotizacion.vendedor_id.is_(None),
                        func.upper(func.trim(Cotizacion.vendedor_nombre)).in_(visible_names),
                    )
                )
            query = query.filter(or_(*seller_filters))

    result = await db.execute(query)
    items_with_quotes = result.all()

    clients = find_clients_for_promotion(
        promotion,
        items_with_quotes,
        only_invoiced=False,  # Include both invoiced and quoted
    )

    # Build pre-filled message for WhatsApp/Email.
    promo_desc = promotion.descripcion_material or promotion.codigo_material or "Material"
    promo_price = f"${promotion.precio_promocion:,.2f}" if promotion.precio_promocion else "Precio especial"
    promo_valid = promotion.valido_hasta.strftime("%d/%m/%Y") if promotion.valido_hasta else "por tiempo limitado"

    wa_message = (
        f"¡Hola! Le informamos que tenemos una promoción especial en "
        f"{promo_desc} a {promo_price} (válido hasta {promo_valid}). "
        f"¿Le interesa? Quedo a sus órdenes."
    )
    email_subject = f"Promoción Especial: {promo_desc} a {promo_price}"
    email_body = (
        f"Estimado cliente,\n\n"
        f"Le informamos que tenemos una promoción especial:\n\n"
        f"  Material: {promo_desc}\n"
        f"  Precio de Promoción: {promo_price}\n"
        f"  Válido hasta: {promo_valid}\n\n"
        f"¡No deje pasar esta oportunidad!\n\nSaludos cordiales."
    )

    # Enrich each client with contact action URLs.
    for client in clients:
        contact = client.get("contacto", {})
        phone = safe_phone_href(contact.get("contacto_preferente"))
        email = contact.get("email")
        client["acciones"] = {
            "whatsapp_url": (
                f"https://wa.me/{phone}?text={__import__('urllib.parse', fromlist=['quote']).quote(wa_message)}"
                if phone
                else None
            ),
            "email_url": (
                f"mailto:{email}?subject={__import__('urllib.parse', fromlist=['quote']).quote(email_subject)}"
                f"&body={__import__('urllib.parse', fromlist=['quote']).quote(email_body)}"
                if email
                else None
            ),
            "telefono": phone,
        }

    return {
        "status": "success",
        "data": {
            "promocion": promotion.to_dict(),
            "mensaje_whatsapp": wa_message,
            "asunto_email": email_subject,
            "clientes": clients,
            "total_clientes": len(clients),
        },
    }
