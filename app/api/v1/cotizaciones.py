from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
import io
import re
import unicodedata
from zoneinfo import ZoneInfo

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import and_, case, delete, func, insert, or_
from sqlalchemy.orm import load_only
from uuid import UUID, uuid4
from typing import Any, Optional

from app.core.database import get_db
from app.core.security import get_current_user, RoleChecker
from app.models.usuario import Usuario
from app.models.cotizacion import Cotizacion
from app.models.cotizacion_detalle import CotizacionComentario, CotizacionItem
from app.models.recordatorio_seguimiento import RecordatorioSeguimiento
from app.models.promocion import Promocion
from app.models.cliente import Cliente
from app.schemas.cotizacion import (
    CotizacionCreate,
    CotizacionCreateManual,
    CotizacionUpdate,
    RecordatorioCreate,
    RecordatorioUpdate,
)
from app.schemas.commercial import ComentarioCreate, ComentarioUpdate
from app.agents.cotizaciones_agent import generate_proposal
from app.services.jerarquia import get_ids_vendedores_visibles
from app.services.commercial_analytics import normalize_contact, normalize_text, promotion_priority
from app.services.client_history_service import build_client_history
from app.core.config import settings

require_admin_or_gerente = RoleChecker(["admin", "gerente"])

router = APIRouter()

MAX_OPERATIONAL_PAGE_SIZE = 100
VALID_QUOTE_STATES = {"all", "total", "concretadas", "pendientes", "vencidas"}
VALID_QUOTE_VIEWS = {"completa", "resumen"}

def _normalize_seller_text(value) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.upper().split())


def _seller_ids_by_name(users: list[Usuario]) -> dict[str, UUID]:
    """Crea un índice de nombres solo cuando la coincidencia es inequívoca."""
    matches: dict[str, Optional[UUID]] = {}
    for user in users:
        normalized_name = _normalize_seller_text(user.nombre_completo)
        if not normalized_name:
            continue
        if normalized_name in matches and matches[normalized_name] != user.id:
            matches[normalized_name] = None
        else:
            matches[normalized_name] = user.id
    return {name: seller_id for name, seller_id in matches.items() if seller_id is not None}


def _business_today() -> date:
    try:
        return datetime.now(ZoneInfo(settings.BUSINESS_TIMEZONE)).date()
    except Exception:
        return date.today()


def _quote_status_conditions(today: date) -> dict[str, object]:
    """Condiciones SQL deterministas que comparten listado y KPI."""
    has_invoice = and_(
        Cotizacion.numero_factura.is_not(None),
        func.length(func.trim(Cotizacion.numero_factura)) > 0,
    )
    explicitly_lost = (
        func.upper(func.trim(func.coalesce(Cotizacion.venta_perdida, ""))) == "SI"
    )
    expired_by_age = and_(
        Cotizacion.fecha_registro.is_not(None),
        Cotizacion.fecha_registro < today - timedelta(days=settings.QUOTE_VALID_DAYS),
    )
    expired = and_(~has_invoice, or_(explicitly_lost, expired_by_age))
    return {
        "total": ~expired,
        "concretadas": has_invoice,
        "pendientes": and_(~has_invoice, ~expired),
        "vencidas": expired,
    }


async def _quote_summary(db: AsyncSession, query, conditions: dict[str, object]) -> dict:
    """Obtiene KPI por SQL sin materializar cotizaciones en Python."""
    metrics = []
    for status_name in ("total", "concretadas", "pendientes", "vencidas"):
        condition = conditions[status_name]
        metrics.extend(
            [
                func.coalesce(
                    func.sum(case((condition, 1), else_=0)), 0
                ).label(f"{status_name}_count"),
                func.coalesce(
                    func.sum(case((condition, Cotizacion.total), else_=0)), 0
                ).label(f"{status_name}_amount"),
                func.coalesce(
                    func.sum(
                        case((condition, Cotizacion.importe_facturado), else_=0)
                    ),
                    0,
                ).label(f"{status_name}_invoiced_amount"),
            ]
        )
    row = (await db.execute(query.with_only_columns(*metrics))).mappings().one()
    return {
        status_name: {
            "count": int(row[f"{status_name}_count"] or 0),
            "amount": float(row[f"{status_name}_amount"] or 0),
            "invoiced_amount": float(row[f"{status_name}_invoiced_amount"] or 0),
        }
        for status_name in ("total", "concretadas", "pendientes", "vencidas")
    }


def _get_scalar_val(obj: Any, attr: str, default: Any = None) -> Any:
    """Obtiene el valor de un atributo de forma segura sin disparar lazy loading síncrono."""
    if obj is None:
        return default
    if hasattr(obj, "__dict__"):
        if attr in obj.__dict__:
            return obj.__dict__[attr]
        try:
            state = getattr(obj, "_sa_instance_state", None)
            if state is not None and attr in state.unloaded:
                return default
        except Exception:
            pass
    try:
        return getattr(obj, attr, default)
    except Exception:
        return default


def serialize_cotizacion(
    c: Cotizacion,
    resolved_vendedor_id: Optional[UUID] = None,
    enrichment: Optional[dict] = None,
    vista: str = "completa",
) -> dict:
    vendedor_id = _get_scalar_val(c, "vendedor_id") or resolved_vendedor_id
    enrichment = enrichment or {}
    contact_data = enrichment.get("datos_contacto") or normalize_contact(_get_scalar_val(c, "datos_contacto"))
    margen_val = _get_scalar_val(c, "margen")
    imp_fac_val = _get_scalar_val(c, "importe_facturado")
    fecha_reg_val = _get_scalar_val(c, "fecha_registro")
    fecha_fac_val = _get_scalar_val(c, "fecha_factura")
    pct_mat_val = _get_scalar_val(c, "porcentaje_materiales")
    pct_imp_val = _get_scalar_val(c, "porcentaje_importe")
    total_val = _get_scalar_val(c, "total")
    cli_nom_raw = _get_scalar_val(c, "cliente_nombre")
    cli_num_raw = _get_scalar_val(c, "numero_cliente")
    num_cot_raw = _get_scalar_val(c, "numero_cotizacion")
    vend_nom_raw = _get_scalar_val(c, "vendedor_nombre") or enrichment.get("vendedor_nombre")

    # Enriquecer cliente de forma inteligente y resiliente
    if enrichment.get("cliente_nombre"):
        resolved_client = enrichment["cliente_nombre"]
    elif cli_nom_raw and cli_nom_raw not in ("Cliente Desconocido", "Cliente sin registrar"):
        resolved_client = cli_nom_raw
    elif contact_data.get("nombre_contacto"):
        resolved_client = contact_data["nombre_contacto"]
    elif cli_num_raw:
        resolved_client = f"Cliente #{cli_num_raw}"
    elif num_cot_raw:
        resolved_client = f"Cotización #{num_cot_raw}"
    else:
        resolved_client = "Cliente"

    data = {
        "id": str(_get_scalar_val(c, "id")),
        "vendedor_id": str(vendedor_id) if vendedor_id else None,
        "vendedor_nombre": vend_nom_raw,
        "vendedor_sin_vincular": vendedor_id is None,
        "cliente_nombre": resolved_client,
        "numero_cliente": cli_num_raw,
        "datos_contacto": contact_data,
        "total": float(total_val) if total_val is not None else 0.0,
        "numero_cotizacion": _get_scalar_val(c, "numero_cotizacion"),
        "fecha_registro": fecha_reg_val.isoformat() if hasattr(fecha_reg_val, "isoformat") else (str(fecha_reg_val) if fecha_reg_val else None),
        "canal": _get_scalar_val(c, "canal"),
        "numero_factura": _get_scalar_val(c, "numero_factura"),
        "fecha_factura": fecha_fac_val.isoformat() if hasattr(fecha_fac_val, "isoformat") else (str(fecha_fac_val) if fecha_fac_val else None),
        "hora_facturacion": _get_scalar_val(c, "hora_facturacion"),
        "margen": float(margen_val) if margen_val is not None else None,
        "grupo_vendedores": _get_scalar_val(c, "grupo_vendedores"),
        "plazo_entrega": _get_scalar_val(c, "plazo_entrega"),
        "importe_facturado": float(imp_fac_val) if imp_fac_val is not None else None,
        "venta_perdida": _get_scalar_val(c, "venta_perdida"),
        "comentarios": _get_scalar_val(c, "comentarios"),
        "comentarios_seguimiento_count": enrichment.get("comentarios_seguimiento_count", 0),
        "tiene_promocion": enrichment.get("tiene_promocion", False),
        "nivel_prioridad": enrichment.get("nivel_prioridad"),
        "promociones_coincidentes": enrichment.get("promociones_coincidentes", []),
        "materiales_cotizados": _get_scalar_val(c, "materiales_cotizados"),
        "materiales_facturados": _get_scalar_val(c, "materiales_facturados"),
        "porcentaje_materiales": float(pct_mat_val) if pct_mat_val is not None else None,
        "porcentaje_importe": float(pct_imp_val) if pct_imp_val is not None else None,
    }
    if vista != "resumen":
        data.update(
            {
                "items": _get_scalar_val(c, "items"),
                "items_detalle": enrichment.get("items_detalle", []),
                "texto_propuesta": _get_scalar_val(c, "texto_propuesta"),
            }
        )
    return data


async def _load_quote_enrichment(
    db: AsyncSession,
    quotes: list[Cotizacion],
    include_items_detail: bool = True,
) -> dict[UUID, dict]:
    if not quotes:
        return {}
    quote_ids = [quote.id for quote in quotes]
    detail_rows = (
        await db.execute(
            select(CotizacionItem).where(CotizacionItem.cotizacion_id.in_(quote_ids))
        )
    ).scalars().all()
    details_by_quote: dict[UUID, list[CotizacionItem]] = defaultdict(list)
    for row in detail_rows:
        details_by_quote[row.cotizacion_id].append(row)

    item_codes = {
        normalize_text(item.codigo_material)
        for item in detail_rows
        if normalize_text(item.codigo_material)
    }
    if not item_codes:
        for q in quotes:
            raw_mat = str(q.__dict__.get("materiales_cotizados") or "") if hasattr(q, "__dict__") else str(getattr(q, "materiales_cotizados", "") or "")
            if raw_mat:
                for token in re.split(r"[,;\s/]+", raw_mat):
                    cleaned = normalize_text(token)
                    if cleaned:
                        item_codes.add(cleaned)
            raw_items = q.__dict__.get("items") if hasattr(q, "__dict__") else getattr(q, "items", None)
            if isinstance(raw_items, list):
                for it in raw_items:
                    if isinstance(it, dict):
                        prod = it.get("codigo_material") or it.get("producto") or it.get("sku") or it.get("descripcion")
                        cleaned = normalize_text(prod)
                        if cleaned:
                            item_codes.add(cleaned)

    promotions = (
        (await db.execute(select(Promocion))).scalars().all()
        if item_codes
        else []
    )

    comment_counts = dict(
        (
            await db.execute(
                select(
                    CotizacionComentario.cotizacion_id,
                    func.count(CotizacionComentario.id),
                )
                .where(CotizacionComentario.cotizacion_id.in_(quote_ids))
                .group_by(CotizacionComentario.cotizacion_id)
            )
        ).all()
    )

    # Cargar contactos desde el catálogo de Clientes para enriquecer cotizaciones
    client_numbers = {q.numero_cliente.strip() for q in quotes if q.numero_cliente and q.numero_cliente.strip()}
    client_names = {q.cliente_nombre.strip() for q in quotes if q.cliente_nombre and q.cliente_nombre.strip()}
    clients_by_number: dict[str, Cliente] = {}
    clients_by_name: dict[str, Cliente] = {}
    if client_numbers or client_names:
        conditions = []
        if client_numbers:
            conditions.append(Cliente.numero_cliente.in_(client_numbers))
        if client_names:
            conditions.append(Cliente.nombre.in_(client_names))
        catalog_clients = (await db.execute(select(Cliente).where(or_(*conditions)))).scalars().all()
        for cl in catalog_clients:
            if cl.numero_cliente and cl.numero_cliente.strip():
                clients_by_number[cl.numero_cliente.strip()] = cl
            if cl.nombre and cl.nombre.strip():
                clients_by_name[cl.nombre.strip()] = cl

    try:
        today = datetime.now(ZoneInfo(settings.BUSINESS_TIMEZONE)).date()
    except Exception:
        today = date.today()

    result = {}
    for quote in quotes:
        items = details_by_quote.get(quote.id, [])
        promo = promotion_priority(
            quote,
            items,
            promotions,
            today,
            settings.QUOTE_VALID_DAYS,
        )

        # Enriquecer datos de contacto con el catálogo (máxima prioridad)
        raw_cli_num = _get_scalar_val(quote, "numero_cliente")
        raw_cli_nom = _get_scalar_val(quote, "cliente_nombre")
        client = clients_by_number.get(raw_cli_num.strip() if raw_cli_num else "") or (clients_by_name.get(raw_cli_nom.strip()) if raw_cli_nom and raw_cli_nom not in ("Cliente Desconocido", "Cliente sin registrar") else None)
        enriched_contact = dict(_get_scalar_val(quote, "datos_contacto") or {})
        resolved_client_name = None
        if client:
            if client.nombre:
                resolved_client_name = client.nombre
            if client.nombre_contacto and client.nombre_contacto.strip():
                enriched_contact["nombre_contacto"] = client.nombre_contacto.strip()
            if client.celular and client.celular.strip():
                enriched_contact["celular"] = client.celular.strip()
            if client.telefono and client.telefono.strip():
                enriched_contact["telefono"] = client.telefono.strip()
            if client.email and client.email.strip():
                enriched_contact["email"] = client.email.strip()
        elif raw_cli_num and (not raw_cli_nom or raw_cli_nom in ("Cliente Desconocido", "Cliente sin registrar")):
            resolved_client_name = f"Cliente #{raw_cli_num}"

        result[quote.id] = {
            **promo,
            "cliente_nombre": resolved_client_name,
            "datos_contacto": normalize_contact(enriched_contact),
            "comentarios_seguimiento_count": comment_counts.get(quote.id, 0),
            "items_detalle": [
                {
                    "id": str(item.id),
                    "codigo_material": item.codigo_material,
                    "descripcion": item.descripcion,
                    "indicador_abcf": item.indicador_abcf,
                    "unidad_medida": item.unidad_medida,
                    "precio_venta": float(item.precio_venta) if item.precio_venta is not None else 0.0,
                    "familia": item.familia,
                    "grupo_materiales": item.grupo_materiales,
                    "cantidad_cotizada": float(item.cantidad_cotizada),
                    "importe_cotizado": float(item.importe_cotizado),
                    "cantidad_facturada": float(item.cantidad_facturada),
                    "importe_facturado": float(item.importe_facturado),
                    "es_promocion": bool(item.es_promocion),
                    "precio_promocion": float(item.precio_promocion) if item.precio_promocion is not None else None,
                }
                for item in items
            ] if include_items_detail else [],
        }
    return result


async def _get_authorized_quote(
    db: AsyncSession,
    current_user: Usuario,
    cotizacion_id: UUID,
) -> Cotizacion:
    cotizacion = (
        await db.execute(select(Cotizacion).where(Cotizacion.id == cotizacion_id))
    ).scalars().first()
    if not cotizacion:
        raise HTTPException(status_code=404, detail="La cotización solicitada no existe.")
    if current_user.rol == "vendedor":
        ids_visibles = await get_ids_vendedores_visibles(db, current_user) or [current_user.id]
        if cotizacion.vendedor_id not in ids_visibles:
            if cotizacion.vendedor_id is not None:
                raise HTTPException(
                    status_code=403,
                    detail="No tienes permiso para consultar esta cotización.",
                )
            visible_users = (
                await db.execute(select(Usuario).where(Usuario.id.in_(ids_visibles)))
            ).scalars().all()
            visible_names = {
                _normalize_seller_text(user.nombre_completo)
                for user in visible_users
                if user.nombre_completo
            }
            if _normalize_seller_text(cotizacion.vendedor_nombre) not in visible_names:
                raise HTTPException(
                    status_code=403,
                    detail="No tienes permiso para consultar esta cotización.",
                )
    return cotizacion

@router.get("/", status_code=status.HTTP_200_OK)
async def list_cotizaciones(
    vendedor_id: Optional[UUID] = None,
    fecha_inicio: Optional[date] = Query(default=None),
    fecha_fin: Optional[date] = Query(default=None),
    busqueda: Optional[str] = Query(default=None, max_length=120),
    total_min: Optional[float] = Query(default=None, ge=0),
    total_max: Optional[float] = Query(default=None, gt=0),
    edad_min_dias: Optional[int] = Query(default=None, ge=0),
    edad_max_dias: Optional[int] = Query(default=None, ge=0),
    sin_vincular: bool = Query(default=False),
    solo_promociones: bool = Query(default=False),
    estado: str = Query(default="all"),
    vista: str = Query(default="completa"),
    orden: str = Query(default="desc"),
    limit: int = Query(default=50, ge=1, le=MAX_OPERATIONAL_PAGE_SIZE),
    offset: int = Query(default=0, ge=0),
    lite: bool = Query(default=False),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Lists all quotes. Salespeople can only list their own.
    Admins and Managers can list all or filter by seller.
    """
    if str(estado).isdigit():
        if edad_max_dias is None:
            edad_max_dias = int(estado)
        estado = "all"

    if estado not in VALID_QUOTE_STATES:
        raise HTTPException(status_code=422, detail="Estado de cotización inválido.")
    if vista not in VALID_QUOTE_VIEWS:
        raise HTTPException(status_code=422, detail="Vista de cotización inválida.")
    if lite and vista == "completa":
        # Compatibilidad con la optimización previa: la vista resumen es más
        # estricta porque tampoco serializa JSON de partidas ni propuestas.
        vista = "resumen"
    if orden not in {"asc", "desc"}:
        raise HTTPException(status_code=422, detail="Orden de cotización inválido.")
    if total_min is not None and total_max is not None and total_min >= total_max:
        raise HTTPException(status_code=422, detail="El rango de total es inválido.")
    if (
        edad_min_dias is not None
        and edad_max_dias is not None
        and edad_min_dias > edad_max_dias
    ):
        raise HTTPException(status_code=422, detail="El rango de antigüedad es inválido.")

    # Filter by vendedor_id based on role
    if current_user.rol == "vendedor":
        # Vendedor: propio + hijos (jerarquia 1 nivel)
        ids_visibles = await get_ids_vendedores_visibles(db, current_user)
        query = select(Cotizacion)
        count_query = select(func.count()).select_from(Cotizacion)
        if ids_visibles is not None:
            visible_users = (
                await db.execute(select(Usuario).where(Usuario.id.in_(ids_visibles)))
            ).scalars().all()
            visible_names = [
                user.nombre_completo.strip().upper()
                for user in visible_users
                if user.nombre_completo and user.nombre_completo.strip()
            ]
            seller_filters = [Cotizacion.vendedor_id.in_(ids_visibles)]
            if visible_names:
                seller_filters.append(
                    and_(
                        Cotizacion.vendedor_id.is_(None),
                        func.upper(func.trim(Cotizacion.vendedor_nombre)).in_(visible_names),
                    )
                )
            seller_condition = or_(*seller_filters)
            query = query.filter(seller_condition)
            count_query = count_query.filter(seller_condition)

    else:
        query = select(Cotizacion)
        count_query = select(func.count()).select_from(Cotizacion)
        if vendedor_id is not None:
            selected_seller = (await db.execute(select(Usuario).where(Usuario.id == vendedor_id))).scalars().first()
            seller_filters = [Cotizacion.vendedor_id == vendedor_id]
            if selected_seller and selected_seller.nombre_completo:
                # Incluye registros históricos donde el Excel guardó el nombre,
                # pero no logró vincular el UUID del usuario al importar.
                seller_filters.append(
                    and_(
                        Cotizacion.vendedor_id.is_(None),
                        func.upper(func.trim(Cotizacion.vendedor_nombre))
                        == selected_seller.nombre_completo.strip().upper(),
                    )
                )
            seller_condition = or_(*seller_filters)
            query = query.filter(seller_condition)
            count_query = count_query.filter(seller_condition)

    if sin_vincular:
        # Los registros con nombre que se resuelve de manera inequívoca siguen
        # siendo visibles por su vendedor; este filtro muestra únicamente los
        # que no tienen un UUID persistido para revisión administrativa.
        query = query.filter(Cotizacion.vendedor_id.is_(None))
        count_query = count_query.filter(Cotizacion.vendedor_id.is_(None))

    # El filtro se resuelve en PostgreSQL para evitar ambigüedades de formato
    # o zona horaria en el navegador.
    if fecha_inicio is not None:
        query = query.filter(Cotizacion.fecha_registro >= fecha_inicio)
        count_query = count_query.filter(Cotizacion.fecha_registro >= fecha_inicio)
    if fecha_fin is not None:
        query = query.filter(Cotizacion.fecha_registro <= fecha_fin)
        count_query = count_query.filter(Cotizacion.fecha_registro <= fecha_fin)

    if busqueda and busqueda.strip():
        search_term = f"%{busqueda.strip()}%"
        search_condition = or_(
            Cotizacion.cliente_nombre.ilike(search_term),
            Cotizacion.numero_cliente.ilike(search_term),
            Cotizacion.numero_cotizacion.ilike(search_term),
            Cotizacion.materiales_cotizados.ilike(search_term),
        )
        query = query.filter(search_condition)
        count_query = count_query.filter(search_condition)

    if total_min is not None:
        query = query.filter(Cotizacion.total >= total_min)
        count_query = count_query.filter(Cotizacion.total >= total_min)
    if total_max is not None:
        query = query.filter(Cotizacion.total < total_max)
        count_query = count_query.filter(Cotizacion.total < total_max)

    # La antigüedad se traduce a fechas de corte en Python, con la zona de
    # negocio, y PostgreSQL sólo compara columnas indexables de fecha.
    today = _business_today()
    if edad_min_dias is not None:
        max_date = today - timedelta(days=edad_min_dias)
        query = query.filter(Cotizacion.fecha_registro <= max_date)
        count_query = count_query.filter(Cotizacion.fecha_registro <= max_date)
    if edad_max_dias is not None:
        min_date = today - timedelta(days=edad_max_dias)
        query = query.filter(Cotizacion.fecha_registro >= min_date)
        count_query = count_query.filter(Cotizacion.fecha_registro >= min_date)

    summary = None
    status_conditions = _quote_status_conditions(today)
    if vista == "resumen":
        summary = await _quote_summary(db, query, status_conditions)

    if estado != "all":
        status_condition = status_conditions[estado]
        query = query.filter(status_condition)
        count_query = count_query.filter(status_condition)

    # Count total quotes
    count_res = await db.execute(count_query)
    total = count_res.scalar_one()

    # List quotes
    if vista == "resumen":
        query = query.options(
            load_only(
                Cotizacion.id,
                Cotizacion.vendedor_id,
                Cotizacion.vendedor_nombre,
                Cotizacion.cliente_nombre,
                Cotizacion.numero_cliente,
                Cotizacion.datos_contacto,
                Cotizacion.total,
                Cotizacion.numero_cotizacion,
                Cotizacion.fecha_registro,
                Cotizacion.canal,
                Cotizacion.numero_factura,
                Cotizacion.fecha_factura,
                Cotizacion.hora_facturacion,
                Cotizacion.margen,
                Cotizacion.grupo_vendedores,
                Cotizacion.plazo_entrega,
                Cotizacion.importe_facturado,
                Cotizacion.venta_perdida,
                Cotizacion.comentarios,
                Cotizacion.materiales_cotizados,
                Cotizacion.materiales_facturados,
                Cotizacion.porcentaje_materiales,
                Cotizacion.porcentaje_importe,
            )
        )
    date_order = (
        Cotizacion.fecha_registro.asc().nullslast()
        if orden == "asc"
        else Cotizacion.fecha_registro.desc().nullslast()
    )
    number_order = (
        Cotizacion.numero_cotizacion.asc().nullslast()
        if orden == "asc"
        else Cotizacion.numero_cotizacion.desc().nullslast()
    )
    query = query.order_by(date_order, number_order).offset(offset).limit(limit)
    res = await db.execute(query)
    cotizaciones = res.scalars().all()

    # Las cotizaciones históricas pueden traer solo el nombre del vendedor desde
    # Excel. Se resuelven al serializar, sin alterar la base, para que los filtros
    # administrativos funcionen también con esos registros.
    users = (await db.execute(select(Usuario))).scalars().all()
    seller_ids_by_name = _seller_ids_by_name(users)
    enrichment = await _load_quote_enrichment(
        db,
        list(cotizaciones),
        include_items_detail=vista != "resumen",
    )
    data = [
        serialize_cotizacion(
            c,
            seller_ids_by_name.get(_normalize_seller_text(c.vendedor_nombre)),
            enrichment.get(c.id),
            vista=vista,
        )
        for c in cotizaciones
    ]

    if solo_promociones:
        data = [d for d in data if d.get("tiene_promocion") is True]
        total = len(data)

    payload = {
        "status": "success",
        "pagination": {
            "limit": limit,
            "offset": offset,
            "total": total
        },
        "data": data
    }
    if summary is not None:
        payload["summary"] = summary
    return payload


@router.get("/historial-cliente", status_code=status.HTTP_200_OK)
async def get_client_history(
    numero_cliente: str = Query(..., min_length=1),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """
    Return aggregate purchase history for a client identified by ``numero_cliente``.

    Salespeople only see operations linked to themselves (or their hierarchy).
    Admins/Managers see all operations for the client.
    """
    if current_user.rol == "soporte":
        raise HTTPException(
            status_code=403,
            detail="El rol de soporte no tiene acceso a la analítica comercial.",
        )

    query = select(Cotizacion).where(
        func.upper(func.trim(Cotizacion.numero_cliente)) == numero_cliente.strip().upper()
    )

    if current_user.rol == "vendedor":
        ids_visibles = await get_ids_vendedores_visibles(db, current_user)
        if ids_visibles is not None:
            visible_users = (
                await db.execute(select(Usuario).where(Usuario.id.in_(ids_visibles)))
            ).scalars().all()
            visible_names = [
                user.nombre_completo.strip().upper()
                for user in visible_users
                if user.nombre_completo and user.nombre_completo.strip()
            ]
            seller_filters = [Cotizacion.vendedor_id.in_(ids_visibles)]
            if visible_names:
                seller_filters.append(
                    and_(
                        Cotizacion.vendedor_id.is_(None),
                        func.upper(func.trim(Cotizacion.vendedor_nombre)).in_(visible_names),
                    )
                )
            query = query.filter(or_(*seller_filters))

    result = await db.execute(query)
    quotes = result.scalars().all()

    try:
        today = datetime.now(ZoneInfo(settings.BUSINESS_TIMEZONE)).date()
    except Exception:
        today = date.today()

    history = build_client_history(
        numero_cliente,
        quotes,
        quote_valid_days=settings.QUOTE_VALID_DAYS,
        today=today,
    )

    return {"status": "success", "data": history}


@router.get("/recordatorios/pendientes", status_code=status.HTTP_200_OK)
async def list_pending_reminders(
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """
    Returns pending follow-up reminders (completado = False).
    Vendedores see their own + hierarchy.
    Admins / Gerentes see all.
    """
    if current_user.rol == "soporte":
        raise HTTPException(
            status_code=403,
            detail="El rol de soporte no tiene acceso a esta funcionalidad.",
        )

    query = select(RecordatorioSeguimiento).where(
        RecordatorioSeguimiento.completado.is_(False)
    )

    if current_user.rol == "vendedor":
        ids_visibles = await get_ids_vendedores_visibles(db, current_user) or [current_user.id]
        query = query.where(RecordatorioSeguimiento.vendedor_id.in_(ids_visibles))

    result = await db.execute(query.order_by(RecordatorioSeguimiento.fecha_programada.asc()))
    reminders = result.scalars().all()

    # Enrich with quote details
    quote_ids = [r.cotizacion_id for r in reminders]
    quotes_map = {}
    if quote_ids:
        quotes_res = await db.execute(select(Cotizacion).where(Cotizacion.id.in_(quote_ids)))
        quotes_map = {q.id: q for q in quotes_res.scalars().all()}

    data = []
    for r in reminders:
        q = quotes_map.get(r.cotizacion_id)
        data.append({
            "id": str(r.id),
            "cotizacion_id": str(r.cotizacion_id),
            "vendedor_id": str(r.vendedor_id),
            "fecha_programada": r.fecha_programada.isoformat(),
            "nota": r.nota,
            "completado": r.completado,
            "creado_en": r.creado_en.isoformat(),
            "cliente_nombre": q.cliente_nombre if q else "Cliente",
            "numero_cliente": q.numero_cliente if q else None,
            "numero_cotizacion": q.numero_cotizacion if q else None,
            "total": float(q.total) if q else 0,
        })

    return {"status": "success", "data": data}


@router.post("/{cotizacion_id}/recordatorio", status_code=status.HTTP_201_CREATED)
async def create_quote_reminder(
    cotizacion_id: UUID,
    payload: RecordatorioCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """
    Creates a new follow-up reminder for a specific quote.
    """
    cotizacion = await _get_authorized_quote(db, current_user, cotizacion_id)

    reminder = RecordatorioSeguimiento(
        cotizacion_id=cotizacion.id,
        vendedor_id=current_user.id,
        fecha_programada=payload.fecha_programada,
        nota=payload.nota.strip() if payload.nota else None,
    )
    db.add(reminder)
    await db.commit()
    await db.refresh(reminder)

    return {
        "status": "success",
        "message": "Recordatorio de seguimiento agendado exitosamente.",
        "data": {
            "id": str(reminder.id),
            "cotizacion_id": str(reminder.cotizacion_id),
            "vendedor_id": str(reminder.vendedor_id),
            "fecha_programada": reminder.fecha_programada.isoformat(),
            "nota": reminder.nota,
            "completado": reminder.completado,
            "creado_en": reminder.creado_en.isoformat(),
        },
    }


@router.patch("/recordatorios/{recordatorio_id}", status_code=status.HTTP_200_OK)
async def update_reminder(
    recordatorio_id: UUID,
    payload: RecordatorioUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """
    Updates or marks a follow-up reminder as completed.
    """
    res = await db.execute(
        select(RecordatorioSeguimiento).where(RecordatorioSeguimiento.id == recordatorio_id)
    )
    reminder = res.scalars().first()
    if not reminder:
        raise HTTPException(status_code=404, detail="El recordatorio solicitado no existe.")

    if current_user.rol == "vendedor" and reminder.vendedor_id != current_user.id:
        ids_visibles = await get_ids_vendedores_visibles(db, current_user) or [current_user.id]
        if reminder.vendedor_id not in ids_visibles:
            raise HTTPException(status_code=403, detail="No tienes permiso para modificar este recordatorio.")

    if payload.fecha_programada is not None:
        reminder.fecha_programada = payload.fecha_programada
    if payload.nota is not None:
        reminder.nota = payload.nota.strip() if payload.nota else None
    if payload.completado is not None:
        reminder.completado = payload.completado
        reminder.completado_en = datetime.now(timezone.utc).replace(tzinfo=None) if payload.completado else None

    await db.commit()
    await db.refresh(reminder)

    return {
        "status": "success",
        "message": "Recordatorio actualizado exitosamente.",
        "data": {
            "id": str(reminder.id),
            "cotizacion_id": str(reminder.cotizacion_id),
            "vendedor_id": str(reminder.vendedor_id),
            "fecha_programada": reminder.fecha_programada.isoformat(),
            "nota": reminder.nota,
            "completado": reminder.completado,
            "completado_en": reminder.completado_en.isoformat() if reminder.completado_en else None,
        },
    }


def _serialize_comment(comment: CotizacionComentario, author: Optional[Usuario] = None) -> dict:
    return {
        "id": str(comment.id),
        "cotizacion_id": str(comment.cotizacion_id),
        "autor_id": str(comment.autor_id) if comment.autor_id else None,
        "autor_nombre": (
            (author.nombre_completo or author.email)
            if author
            else "Usuario eliminado"
        ),
        "comentario": comment.comentario,
        "creado_en": comment.creado_en.isoformat(),
        "editado_en": comment.editado_en.isoformat() if comment.editado_en else None,
    }


@router.get("/{cotizacion_id}/comentarios", status_code=status.HTTP_200_OK)
async def list_quote_comments(
    cotizacion_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    await _get_authorized_quote(db, current_user, cotizacion_id)
    rows = (
        await db.execute(
            select(CotizacionComentario, Usuario)
            .outerjoin(Usuario, Usuario.id == CotizacionComentario.autor_id)
            .where(CotizacionComentario.cotizacion_id == cotizacion_id)
            .order_by(CotizacionComentario.creado_en.asc())
        )
    ).all()
    return {
        "status": "success",
        "data": [_serialize_comment(comment, author) for comment, author in rows],
    }


@router.post(
    "/{cotizacion_id}/comentarios",
    status_code=status.HTTP_201_CREATED,
)
async def create_quote_comment(
    cotizacion_id: UUID,
    payload: ComentarioCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    if current_user.rol not in ("admin", "gerente", "vendedor"):
        raise HTTPException(
            status_code=403,
            detail="Tu rol no puede agregar comentarios de seguimiento.",
        )
    await _get_authorized_quote(db, current_user, cotizacion_id)
    text = payload.comentario.strip()
    if not text:
        raise HTTPException(status_code=400, detail="El comentario no puede estar vacío.")
    comment = CotizacionComentario(
        cotizacion_id=cotizacion_id,
        autor_id=current_user.id,
        comentario=text,
    )
    db.add(comment)
    await db.commit()
    await db.refresh(comment)
    return {
        "status": "success",
        "message": "Comentario agregado.",
        "data": _serialize_comment(comment, current_user),
    }


@router.put(
    "/{cotizacion_id}/comentarios/{comentario_id}",
    status_code=status.HTTP_200_OK,
)
async def update_quote_comment(
    cotizacion_id: UUID,
    comentario_id: UUID,
    payload: ComentarioUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    if current_user.rol not in ("admin", "gerente", "vendedor"):
        raise HTTPException(
            status_code=403,
            detail="Tu rol no puede editar comentarios de seguimiento.",
        )
    await _get_authorized_quote(db, current_user, cotizacion_id)
    comment = (
        await db.execute(
            select(CotizacionComentario).where(
                CotizacionComentario.id == comentario_id,
                CotizacionComentario.cotizacion_id == cotizacion_id,
            )
        )
    ).scalars().first()
    if not comment:
        raise HTTPException(status_code=404, detail="El comentario no existe.")
    if current_user.rol not in ("admin", "gerente") and comment.autor_id != current_user.id:
        raise HTTPException(status_code=403, detail="No puedes editar comentarios de otro usuario.")
    text = payload.comentario.strip()
    if not text:
        raise HTTPException(status_code=400, detail="El comentario no puede estar vacío.")
    comment.comentario = text
    comment.editado_en = datetime.now(timezone.utc).replace(tzinfo=None)
    await db.commit()
    await db.refresh(comment)
    author = (
        await db.execute(select(Usuario).where(Usuario.id == comment.autor_id))
    ).scalars().first()
    return {
        "status": "success",
        "message": "Comentario actualizado.",
        "data": _serialize_comment(comment, author),
    }


def _excel_identifier(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _excel_number(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        raise ValueError(f"valor numérico inválido: {value}")


QUOTE_IMPORT_REQUIRED_HEADERS = {
    "fecha_registro": "FECHA DE REGISTRO",
    "organizacion_ventas": "ORGANIZACION DE VENTAS",
    "numero_cotizacion": "NUMERO DE COTIZACION",
    "vendedor_codigo": "VENDEDOR",
    "vendedor_nombre": "NOMBRE DEL VENDEDOR",
    "numero_cliente": "NUMERO DEL CLIENTE",
    "cliente_nombre": "NOMBRE DEL CLIENTE",
    "telefono": "NUMERO DE TELEFONO",
    "celular": "NUMERO DE CELULAR",
    "email": "DIRECCION CORREO ELECTRONICO",
    "numero_factura": "NUMERO DE FACTURA",
    "fecha_factura": "FECHA DE FACTURA",
    "importe_cotizado": "IMPORTE COTIZADO C/IVA",
    "importe_facturado": "IMPORTE FACTURADO C/IVA",
    "porcentaje_importe": "PORCENTAJE DE IMPORTE",
    "materiales_cotizados": "MATERIALES COTIZADOS",
    "materiales_facturados": "MATERIALES FACTURADOS",
    "porcentaje_materiales": "PORCENTAJE DE MATERIALES",
}
QUOTE_IMPORT_CHANNEL_HEADER = "TIPO DE ENTREGA"
QUOTE_IMPORT_LEGACY_CHANNEL_HEADER = "CANAL"

DETAIL_IMPORT_REQUIRED_HEADERS = {
    "numero_cotizacion": "NUMERO DE COTIZACION",
    "codigo_material": "CODIGO MATERIAL",
    "descripcion": "DESCRIPCION",
    "familia": "FAMILIA",
    "grupo_materiales": "GRUPO DE MATERIALES",
    "cantidad_cotizada": "CANTIDAD COTIZADA",
    "importe_cotizado": "IMPORTE COTIZADO",
    "cantidad_facturada": "CANTIDAD FACTURADA",
    "importe_facturado": "IMPORTE FACTURADO",
}

MULTISHEET_VENTAS_REQUIRED_HEADERS = {
    "fecha_factura": "FECHA DE FACTURA",
    "numero_cliente": "NUMERO DEL CLIENTE",
    "plazo_entrega": "PLAZO DE ENTREGA",
    "nombre_cliente": "NOMBRE DEL CLIENTE",
    "folio_cotizacion": "FOLIO COTIZACION",
    "folio_factura": "FOLIO FACTURA",
    "hora_facturacion": "HORA DE FACTURACION",
    "margen": "MARGEN",
    "grupo_vendedores": "GRUPO DE VENDEDORES",
    "nombre_vendedor": "NOMBRE DEL VENDEDOR",
    "canal_distribucion": "CANAL DE DISTRIBUCION",
}

MULTISHEET_COTIZACIONES_REQUIRED_HEADERS = {
    "fecha_registro": "FECHA DE REGISTRO",
    "organizacion_ventas": "ORGANIZACION DE VENTAS",
    "numero_cotizacion": "NUMERO DE COTIZACION",
    "indicador_abcf": "INDICADOR ABC+FRECUENCIA DE VENTA",
    "codigo_material": "CODIGO DE MATERIAL",
    "descripcion_material": "DESCRIPCION DEL MATERIAL",
    "unidad_medida": "UNIDAD DE MEDIDA",
    "precio_venta": "PRECIO DE VENTA",
}


def _resolve_sheet_columns(worksheet) -> tuple[dict[str, Optional[int]], str]:
    """
    Analiza una hoja de cálculo y determina si es una hoja de cabecera (HEADER),
    una hoja de detalle de partidas (DETAIL) o mixta (BOTH), resolviendo los índices de columnas.
    """
    headers = {}
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for idx, val in enumerate(header_row):
        if val is not None:
            norm = normalize_text(str(val))
            if norm:
                headers[norm] = idx

    def find_idx(candidates: list[str]) -> Optional[int]:
        # 1. Coincidencia exacta normalizada
        for c in candidates:
            c_norm = normalize_text(c)
            if c_norm in headers:
                return headers[c_norm]
        # 2. Coincidencia por subcadena específica (mínimo 4 caracteres para evitar falsos positivos)
        for c in candidates:
            c_norm = normalize_text(c)
            if len(c_norm) >= 4:
                for h_name, h_idx in headers.items():
                    if c_norm in h_name:
                        return h_idx
        return None

    col_map = {
        "numero_cotizacion": find_idx(["NUMERO DE COTIZACION", "NUMERO COTIZACION", "FOLIO COTIZACION", "FOLIO DE COTIZACION", "COTIZACION", "NUMERO_COTIZACION", "DOC. VENTAS", "DOCUMENTO VENTAS"]),
        "codigo_material": find_idx(["CODIGO DE MATERIAL", "CODIGO MATERIAL", "CODIGO DEL MATERIAL", "CODIGO_MATERIAL", "CODIGO", "SKU", "CLAVE MATERIAL", "ARTICULO", "MATERIAL"]),
        "descripcion": find_idx(["DESCRIPCION DEL MATERIAL", "DESCRIPCION DE MATERIAL", "DESCRIPCION MATERIAL", "DESCRIPCION", "CONCEPTO", "NOMBRE MATERIAL"]),
        "precio": find_idx(["PRECIO DE VENTA", "PRECIO UNITARIO", "PRECIO", "IMPORTE COTIZADO", "IMPORTE", "VALOR NETO", "TOTAL"]),
        "unidad_medida": find_idx(["UNIDAD DE MEDIDA", "UNIDAD DE MEDIDA BASE", "UNIDAD MEDIDA", "UMB", "UDM", "UNIDAD"]),
        "indicador_abcf": find_idx(["INDICADOR ABC+FRECUENCIA DE VENTA", "ABC+F", "ABCF", "INDICADOR ABC", "ABC"]),
        "cliente_nombre": find_idx(["NOMBRE DEL CLIENTE", "NOMBRE DE CLIENTE", "CLIENTE NOMBRE", "NOMBRE CLIENTE", "CLIENTE", "SOLICITANTE", "NOMBRE SOLICITANTE", "RAZON SOCIAL", "EMPRESA", "CUENTA", "DESTINATARIO", "CLIENTE/PROSPECTO", "PROSPECTO"]),
        "numero_cliente": find_idx(["NUMERO DEL CLIENTE", "NUMERO DE CLIENTE", "NUM CLIENTE", "CODIGO CLIENTE", "CODIGO DEL CLIENTE", "CLIENTE ID", "ID CLIENTE", "COD. CLIENTE", "NO. CLIENTE", "SOLICITANTE ID", "NO CLIENTE", "NO. DE CLIENTE", "CVE CLIENTE"]),
        "vendedor_nombre": find_idx(["NOMBRE DEL VENDEDOR", "NOMBRE DE VENDEDOR", "VENDEDOR NOMBRE", "NOMBRE VENDEDOR", "ASESOR NOMBRE", "NOMBRE ASESOR", "NOMBRE DEL ASESOR", "EJECUTIVO", "AGENTE", "RESPONSABLE", "ASESOR", "VENDEDOR"]),
        "vendedor_codigo": find_idx(["NUMERO DE VENDEDOR", "NUMERO VENDEDOR", "CODIGO VENDEDOR", "CODIGO DE VENDEDOR", "CLAVE VENDEDOR", "NUM ASESOR", "CVE ASESOR", "CVE VENDEDOR", "USR VENDEDOR", "VENDEDOR ID", "ID VENDEDOR", "VENDEDOR"]),
        "grupo_vendedores": find_idx(["GRUPO DE VENDEDORES", "GRUPO VENDEDORES", "GRUPO"]),
        "canal": find_idx(["CANAL DE DISTRIBUCION", "CANAL"]),
        "plazo_entrega": find_idx(["PLAZO DE ENTREGA", "TIPO DE ENTREGA", "ENTREGA"]),
        "numero_factura": find_idx(["FOLIO DE LA FACTURA", "FOLIO DE FACTURA", "FOLIO FACTURA", "NUMERO DE FACTURA", "NUMERO FACTURA", "FACTURA", "FOLIO_FACTURA"]),
        "fecha_factura": find_idx(["FECHA DE LA FACTURA", "FECHA DE FACTURA", "FECHA FACTURA"]),
        "hora_facturacion": find_idx(["HORA DE LA FACTURA", "HORA DE FACTURACION", "HORA FACTURA", "HORA FACTURACION", "HORA"]),
        "margen": find_idx(["MARGEN"]),
        "fecha_registro": find_idx(["FECHA DE REGISTRO", "FECHA REGISTRO", "FECHA"]),
        "organizacion_ventas": find_idx(["ORGANIZACION DE VENTAS", "ORG VENTAS", "CENTRO", "SUCURSAL"]),
        "cantidad_facturada": find_idx(["CANTIDAD FACTURADA UMB", "CANTIDAD FACTURADA", "CANTIDAD FACT", "CANTIDAD ENTREGADA"]),
        "importe_facturado": find_idx(["IMPORTE CON IVA", "IMPORTE FACTURADO C/IVA", "IMPORTE FACTURADO CON IVA", "IMPORTE FACTURADO", "IMPORTE FACT"]),
        "telefono": find_idx(["NUMERO DE TELEFONO", "TELEFONO"]),
        "celular": find_idx(["NUMERO DE CELULAR", "CELULAR"]),
        "email": find_idx(["DIRECCION CORREO ELECTRONICO", "CORREO ELECTRONICO", "EMAIL", "CORREO"]),
    }

    has_quote_num = col_map["numero_cotizacion"] is not None
    has_material = col_map["codigo_material"] is not None
    has_client = col_map["cliente_nombre"] is not None or col_map["numero_cliente"] is not None
    has_invoice = col_map["numero_factura"] is not None
    has_seller = col_map["vendedor_nombre"] is not None

    if has_quote_num and has_material and not has_client and not has_invoice and not has_seller:
        sheet_type = "DETAIL"
    elif has_quote_num and (has_client or has_invoice or has_seller) and not has_material:
        sheet_type = "HEADER"
    elif has_quote_num and has_material and (has_client or has_invoice or has_seller):
        sheet_type = "BOTH"
    else:
        sheet_type = "UNKNOWN"

    return col_map, sheet_type


def _detect_multi_sheet_worksheets(workbook) -> Optional[tuple[any, dict[str, Optional[int]], any, dict[str, Optional[int]]]]:
    """
    Detecta automáticamente la hoja de cabecera/ventas y la hoja de detalle de materiales
    analizando columnas y títulos de las hojas.
    """
    visible_sheets = [ws for ws in workbook.worksheets if ws.sheet_state != "hidden"]
    if len(visible_sheets) < 2:
        return None

    header_candidates = []
    detail_candidates = []

    for ws in visible_sheets:
        title_norm = normalize_text(ws.title)
        if "INSTRUCCION" in title_norm or "GUIA" in title_norm or "INSTRUCTION" in title_norm:
            continue

        col_map, sheet_type = _resolve_sheet_columns(ws)
        if col_map["numero_cotizacion"] is None:
            continue

        if sheet_type == "HEADER" or "VENTA" in title_norm:
            header_candidates.append((ws, col_map))
        elif sheet_type == "DETAIL" or any(k in title_norm for k in ("COTIZACION", "DETALLE", "PARTIDA", "DESCRIPCION", "MATERIAL", "PRODUCTO")):
            detail_candidates.append((ws, col_map))
        elif sheet_type == "BOTH":
            # Puede fungir de cabecera si otra hoja es detalle, o de detalle si otra es cabecera
            header_candidates.append((ws, col_map))
            detail_candidates.append((ws, col_map))

    for h_ws, h_cols in header_candidates:
        for d_ws, d_cols in detail_candidates:
            if h_ws != d_ws:
                return h_ws, h_cols, d_ws, d_cols

    return None


def is_multi_sheet_quote_format(workbook) -> bool:
    """Verifica si el archivo contiene dos hojas estructuradas (cabecera/ventas y detalle/cotizaciones)."""
    return _detect_multi_sheet_worksheets(workbook) is not None


def generate_excel_template_bytes() -> bytes:
    """Genera la plantilla oficial descargable en Excel con pestañas Ventas, Cotizaciones e Instrucciones."""
    wb = openpyxl.Workbook()

    # Hoja 1: Ventas
    ws_v = wb.active
    ws_v.title = "Ventas"
    ws_v.append([
        "Fecha de la Factura",
        "Numero del Cliente",
        "Plazo de Entrega",
        "Nombre del Cliente",
        "Folio Cotizacion",
        "Folio de la Factura",
        "Hora de la Factura",
        "Codigo de Material",
        "Descripcion del Material",
        "Cantidad Facturada UMB",
        "Importe con IVA",
        "Margen",
        "Numero de Vendedor",
        "Nombre del Vendedor",
        "Indicador ABC+Frecuencia de Venta",
        "Canal de Distribucion",
    ])
    ws_v.append([
        "2026-01-15",
        "400191",
        "ENTREGA INMEDIATA",
        "PRODIVERSO CASA KURODA",
        "416662481",
        "1325607092",
        "09:52:16",
        "VMVP25",
        "VAL PIE PICHANCHA 25MM VAL-MEX",
        1,
        158.02,
        33.52,
        "C82",
        "Aaron Emigdio Lechuga",
        "B3",
        "01",
    ])
    ws_v.append([
        "2026-01-15",
        "400200",
        "SOBREPEDIDO",
        "CONSTRUCTORA DEL NORTE SA DE CV",
        "416662488",
        "1325607094",
        "10:03:44",
        "TUBOPVC2",
        "TUBO PVC HIDRAULICO 2 PULG",
        2,
        488.36,
        24.40,
        "C94",
        "Jesus Manuel Chavez",
        "A2",
        "01",
    ])

    # Hoja 2: Cotizaciones
    ws_c = wb.create_sheet(title="Cotizaciones")
    ws_c.append([
        "Fecha de Registro",
        "Organizacion de Ventas",
        "Numero de Cotizacion",
        "Indicador ABC+Frecuencia de Venta",
        "Codigo de Material",
        "Descripcion del Material",
        "Unidad de Medida",
        "Precio de Venta",
        "Nombre del Cliente",
        "Numero del Cliente",
        "Nombre del Vendedor",
        "Vendedor",
        "Canal",
    ])
    ws_c.append([
        "2026-01-02",
        "MK01",
        "416662481",
        "C6",
        "VMVP25",
        "VAL PIE PICHANCHA 25MM VAL-MEX",
        "PZA",
        136.21,
        "PRODIVERSO CASA KURODA",
        "400191",
        "Aaron Emigdio Lechuga",
        "C82",
        "01",
    ])
    ws_c.append([
        "2026-01-02",
        "MK01",
        "416662481",
        "A1",
        "TUBOPVC2",
        "TUBO PVC HIDRAULICO 2 PULG",
        "TRAMO",
        210.50,
        "PRODIVERSO CASA KURODA",
        "400191",
        "Aaron Emigdio Lechuga",
        "C82",
        "01",
    ])
    ws_c.append([
        "2026-01-02",
        "MK01",
        "416662488",
        "D6",
        "CSIAZU23NAV32",
        "INSERTO AZUCENA 20X30 NAVIA BEIGE PZA",
        "PZA",
        42.24,
        "CONSTRUCTORA DEL NORTE SA DE CV",
        "400200",
        "Jesus Manuel Chavez",
        "C94",
        "01",
    ])

    # Hoja 3: Instrucciones
    ws_i = wb.create_sheet(title="Instrucciones")
    ws_i.append(["ESTRUCTURA Y GUÍA DE IMPORTACIÓN MULTI-HOJA"])
    ws_i.append(["1. El archivo debe contener dos hojas principales: 'Ventas' y 'Cotizaciones'."])
    ws_i.append(["2. Pestaña 'Ventas': Almacena la facturación emitida, clientes, margen real y canal."])
    ws_i.append(["3. Pestaña 'Cotizaciones': Almacena el desglose por partida / concepto (SKU, descripción, precio unitario)."])
    ws_i.append(["4. La columna 'Folio Cotizacion' en Ventas enlaza automáticamente con 'Numero de Cotizacion' en Cotizaciones."])
    ws_i.append(["5. Las cotizaciones sin factura en Ventas se registran como cotizaciones vivas / en seguimiento."])
    ws_i.append(["6. Los SKUs que coincidan con el catálogo de Promociones se clasificarán como ventas de promoción."])
    ws_i.append(["7. La pestaña 'Cotizaciones' puede incluir opcionalmente 'Nombre del Cliente', 'Numero del Cliente', 'Nombre del Vendedor', 'Vendedor' y 'Canal' para cotizaciones en seguimiento no facturadas."])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



def _quote_import_column_indices(worksheet) -> dict[str, int]:
    """Resuelve el layout por encabezado y prioriza el canal del formato nuevo."""
    headers = {}
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for index, val in enumerate(header_row):
        header = normalize_text(val)
        if header:
            headers[header] = index
    missing = [
        label
        for label in QUOTE_IMPORT_REQUIRED_HEADERS.values()
        if label not in headers
    ]
    channel_header = (
        QUOTE_IMPORT_CHANNEL_HEADER
        if QUOTE_IMPORT_CHANNEL_HEADER in headers
        else QUOTE_IMPORT_LEGACY_CHANNEL_HEADER
    )
    if channel_header not in headers:
        missing.append(
            f"{QUOTE_IMPORT_CHANNEL_HEADER} (o {QUOTE_IMPORT_LEGACY_CHANNEL_HEADER} legado)"
        )
    if missing:
        raise ValueError("Faltan columnas requeridas: " + ", ".join(missing))

    indices = {
        field: headers[label]
        for field, label in QUOTE_IMPORT_REQUIRED_HEADERS.items()
    }
    indices["canal"] = headers[channel_header]
    return indices


def _find_all_quote_worksheets(workbook) -> list[tuple[any, dict[str, int]]]:
    """Busca todas las hojas válidas que contienen la estructura de cotizaciones (ej. múltiples meses)."""
    valid_sheets = []
    last_error = None
    for ws in workbook.worksheets:
        if ws.sheet_state == "hidden":
            continue
        try:
            indices = _quote_import_column_indices(ws)
            valid_sheets.append((ws, indices))
        except ValueError as err:
            last_error = err

    if valid_sheets:
        return valid_sheets
    if last_error:
        raise last_error
    raise ValueError("No se encontraron hojas válidas en el archivo Excel.")


def _find_quote_worksheet(workbook) -> tuple[any, dict[str, int]]:
    """Busca automáticamente la primera hoja que contiene la estructura completa de cotizaciones."""
    return _find_all_quote_worksheets(workbook)[0]


def _find_detail_worksheet(workbook) -> tuple[any, dict[str, int]]:
    """Busca automáticamente la hoja que contiene el detalle de materiales SKU."""
    last_error = None
    for ws in workbook.worksheets:
        if ws.sheet_state == "hidden":
            continue
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = {
            normalize_text(val): index
            for index, val in enumerate(header_row)
            if val is not None
        }
        missing = [label for label in DETAIL_IMPORT_REQUIRED_HEADERS.values() if label not in headers]
        if not missing:
            return ws, headers
        last_error = f"Faltan columnas requeridas: {', '.join(missing)}"

    raise ValueError(last_error or "No se encontraron hojas válidas para detalle de materiales.")


def _merge_quote_fields(existing: Cotizacion, incoming: dict) -> dict:
    """
    Mezcla inteligentemente los campos importados con los existentes para
    evitar sobreescribir clientes, contactos o asesores con valores vacíos o 'Cliente Desconocido'.
    """
    merged = dict(incoming)

    # 1. Preservar cliente si el nuevo viene como 'Cliente Desconocido' o None
    existing_client = getattr(existing, "cliente_nombre", None)
    incoming_client = incoming.get("cliente_nombre")
    if (not incoming_client or incoming_client in ("Cliente Desconocido", "Cliente sin registrar")) and existing_client and existing_client not in ("Cliente Desconocido", "Cliente sin registrar"):
        merged["cliente_nombre"] = existing_client

    existing_client_num = getattr(existing, "numero_cliente", None)
    if not incoming.get("numero_cliente") and existing_client_num:
        merged["numero_cliente"] = existing_client_num

    # Si cliente_nombre sigue siendo vacío o Desconocido pero hay numero_cliente
    if (not merged.get("cliente_nombre") or merged.get("cliente_nombre") in ("Cliente Desconocido", "Cliente sin registrar")) and merged.get("numero_cliente"):
        merged["cliente_nombre"] = f"Cliente #{merged['numero_cliente']}"

    # 2. Preservar asesor
    existing_vend_nom = getattr(existing, "vendedor_nombre", None)
    if not incoming.get("vendedor_nombre") and existing_vend_nom:
        merged["vendedor_nombre"] = existing_vend_nom

    existing_vend_id = getattr(existing, "vendedor_id", None)
    if not incoming.get("vendedor_id") and existing_vend_id:
        merged["vendedor_id"] = existing_vend_id

    existing_grp_vend = getattr(existing, "grupo_vendedores", None)
    if not incoming.get("grupo_vendedores") and existing_grp_vend:
        merged["grupo_vendedores"] = existing_grp_vend

    # 3. Preservar canal y entrega si el nuevo viene vacío
    existing_canal = getattr(existing, "canal", None)
    if not incoming.get("canal") and existing_canal:
        merged["canal"] = existing_canal

    existing_plazo = getattr(existing, "plazo_entrega", None)
    if not incoming.get("plazo_entrega") and existing_plazo:
        merged["plazo_entrega"] = existing_plazo

    existing_org = getattr(existing, "organizacion_ventas", None)
    if not incoming.get("organizacion_ventas") and existing_org:
        merged["organizacion_ventas"] = existing_org

    # 4. Preservar datos de contacto si los nuevos vienen vacíos
    existing_contacts = getattr(existing, "datos_contacto", None) or {}
    incoming_contacts = incoming.get("datos_contacto") or {}
    has_incoming_contact = any(bool(v) for v in incoming_contacts.values())
    if not has_incoming_contact and existing_contacts:
        merged["datos_contacto"] = existing_contacts
    elif incoming_contacts and existing_contacts:
        merged_contacts = dict(existing_contacts)
        for k, v in incoming_contacts.items():
            if v:
                merged_contacts[k] = v
        merged["datos_contacto"] = merged_contacts

    return merged


def _apply_imported_quote_values(cotizacion: Cotizacion, values: dict) -> Cotizacion:
    """Actualiza únicamente campos provenientes del Excel y conserva seguimiento."""
    for field, value in values.items():
        setattr(cotizacion, field, value)
    return cotizacion


def _stale_imported_quote_ids(
    existing_quotes: list[Cotizacion],
    retained_ids: set[UUID],
) -> set[UUID]:
    """Sólo las filas importadas con folio pueden ser removidas al reconciliar."""
    return {
        quote.id
        for quote in existing_quotes
        if _excel_identifier(quote.numero_cotizacion)
        and quote.id not in retained_ids
    }


@router.post("/detalle-materiales/upload", status_code=status.HTTP_201_CREATED)
async def upload_quote_material_detail(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(require_admin_or_gerente),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El detalle debe ser un archivo .xlsx.")
    contents = await file.read()
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
        worksheet, headers = _find_detail_worksheet(workbook)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {exc}")

    quote_rows = (await db.execute(select(Cotizacion))).scalars().all()
    quotes_by_number: dict[str, list[Cotizacion]] = defaultdict(list)
    for quote in quote_rows:
        number = _excel_identifier(quote.numero_cotizacion)
        if number:
            quotes_by_number[number].append(quote)

    accepted: list[CotizacionItem] = []
    rejected = []
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not any(value not in (None, "") for value in row):
            continue
        try:
            quote_number = _excel_identifier(row[headers[DETAIL_IMPORT_REQUIRED_HEADERS["numero_cotizacion"]]])
            matches = quotes_by_number.get(quote_number or "", [])
            if len(matches) != 1:
                reason = (
                    "cotización inexistente"
                    if not matches
                    else "número de cotización ambiguo"
                )
                rejected.append({"fila": row_number, "cotizacion": quote_number, "motivo": reason})
                continue
            code = _excel_identifier(row[headers[DETAIL_IMPORT_REQUIRED_HEADERS["codigo_material"]]])
            if not code:
                rejected.append(
                    {"fila": row_number, "cotizacion": quote_number, "motivo": "SKU vacío"}
                )
                continue
            accepted.append(
                CotizacionItem(
                    cotizacion_id=matches[0].id,
                    codigo_material=code,
                    descripcion=_excel_identifier(row[headers[DETAIL_IMPORT_REQUIRED_HEADERS["descripcion"]]]),
                    familia=_excel_identifier(row[headers[DETAIL_IMPORT_REQUIRED_HEADERS["familia"]]]),
                    grupo_materiales=_excel_identifier(
                        row[headers[DETAIL_IMPORT_REQUIRED_HEADERS["grupo_materiales"]]]
                    ),
                    cantidad_cotizada=_excel_number(
                        row[headers[DETAIL_IMPORT_REQUIRED_HEADERS["cantidad_cotizada"]]]
                    ),
                    importe_cotizado=_excel_number(
                        row[headers[DETAIL_IMPORT_REQUIRED_HEADERS["importe_cotizado"]]]
                    ),
                    cantidad_facturada=_excel_number(
                        row[headers[DETAIL_IMPORT_REQUIRED_HEADERS["cantidad_facturada"]]]
                    ),
                    importe_facturado=_excel_number(
                        row[headers[DETAIL_IMPORT_REQUIRED_HEADERS["importe_facturado"]]]
                    ),
                )
            )
        except Exception as exc:
            rejected.append(
                {"fila": row_number, "cotizacion": None, "motivo": str(exc)}
            )

    if not accepted:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Ninguna partida válida; el detalle existente no fue modificado.",
                "rechazadas": rejected[:100],
            },
        )
    accepted_quote_ids = {item.cotizacion_id for item in accepted}
    await db.execute(
        delete(CotizacionItem).where(
            CotizacionItem.cotizacion_id.in_(accepted_quote_ids)
        )
    )
    db.add_all(accepted)
    await db.commit()
    return {
        "status": "success",
        "message": "Detalle de materiales reemplazado correctamente.",
        "aceptadas": len(accepted),
        "rechazadas": len(rejected),
        "detalle_rechazos": rejected[:100],
    }


@router.get("/plantilla", status_code=status.HTTP_200_OK)
async def download_quote_template():
    """Descarga la plantilla oficial multi-hoja (.xlsx) para importación de Cotizaciones y Ventas."""
    content = generate_excel_template_bytes()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=Plantilla_Cotizaciones_Ventas_Kuroda.xlsx"
        },
    )


@router.post("/upload", status_code=status.HTTP_201_CREATED)
async def upload_cotizaciones(
    file: UploadFile = File(...),
    current_user: Usuario = Depends(require_admin_or_gerente),
):
    if not file.filename or not file.filename.lower().endswith((".xls", ".xlsx")):
        raise HTTPException(
            status_code=400,
            detail="Formato de archivo inválido. Sube un archivo de Excel.",
        )
    contents = await file.read()
    error_msg = await process_excel_background(contents, current_user.id)
    if error_msg:
        raise HTTPException(status_code=500, detail=error_msg)
    return {"message": "El archivo se ha procesado exitosamente."}


@router.get("/{cotizacion_id}", status_code=status.HTTP_200_OK)
async def get_cotizacion(
    cotizacion_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Retrieves detailed information of a single quote. Enforces ownership check for salespeople."""
    cotizacion = await _get_authorized_quote(db, current_user, cotizacion_id)
    enrichment = await _load_quote_enrichment(db, [cotizacion])

    return {
        "status": "success",
        "data": serialize_cotizacion(cotizacion, enrichment=enrichment.get(cotizacion.id))
    }

@router.post("/manual", status_code=status.HTTP_201_CREATED)
async def create_cotizacion_manual(
    quote_in: CotizacionCreateManual,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Creates a quote manually with preset total and proposal text."""
    # Vendedores can only create quotes for themselves.
    # Managers/Admins can assign a quote to any seller (using current_user as fallback)
    new_quote = Cotizacion(
        vendedor_id=current_user.id,
        cliente_nombre=quote_in.cliente_nombre,
        datos_contacto=quote_in.datos_contacto,
        items=quote_in.items,
        total=quote_in.total,
        texto_propuesta=quote_in.texto_propuesta
    )

    db.add(new_quote)
    await db.commit()
    await db.refresh(new_quote)

    return {
        "status": "success",
        "message": "Cotización creada manualmente con éxito.",
        "data": serialize_cotizacion(new_quote)
    }

@router.post("/generate", status_code=status.HTTP_201_CREATED)
async def generate_cotizacion_agente(
    quote_in: CotizacionCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Creates a quote automatically. Invokes the Cotizaciones Agent to
    mathematically compute the total and draft a formal commercial proposal.
    """
    req_adicionales = quote_in.requerimientos_adicionales or ""
    
    # Trigger AI generation
    generated = await generate_proposal(
        cliente_nombre=quote_in.cliente_nombre,
        items=quote_in.items,
        requerimientos_adicionales=req_adicionales
    )

    new_quote = Cotizacion(
        vendedor_id=current_user.id,
        cliente_nombre=quote_in.cliente_nombre,
        datos_contacto=quote_in.datos_contacto,
        items=generated["items_procesados"],
        total=generated["total"],
        texto_propuesta=generated["texto_propuesta"]
    )

    db.add(new_quote)
    await db.commit()
    await db.refresh(new_quote)

    return {
        "status": "success",
        "message": "Cotización generada exitosamente por el Agente de Cotizaciones.",
        "data": serialize_cotizacion(new_quote)
    }

@router.put("/{cotizacion_id}", status_code=status.HTTP_200_OK)
async def update_cotizacion(
    cotizacion_id: UUID,
    quote_in: CotizacionUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Updates a quote. Salespeople can only update their own quotes.
    Admins and Managers can update any quote.
    """
    cotizacion = await _get_authorized_quote(db, current_user, cotizacion_id)

    # Update fields
    update_data = quote_in.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(cotizacion, key, value)

    # Automatically set fecha_factura if numero_factura is set
    if "numero_factura" in update_data and update_data["numero_factura"]:
        from datetime import date
        if not cotizacion.fecha_factura:
            cotizacion.fecha_factura = date.today()

    await db.commit()
    await db.refresh(cotizacion)
    enrichment = await _load_quote_enrichment(db, [cotizacion])

    return {
        "status": "success",
        "message": "Cotización actualizada con éxito.",
        "data": serialize_cotizacion(
            cotizacion,
            enrichment=enrichment.get(cotizacion.id),
        )
    }


async def process_excel_background(contents: bytes, uploaded_by_id: UUID):
    from app.core.database import SessionLocal
    from app.models.cotizacion import Cotizacion
    from app.models.cotizacion_detalle import CotizacionItem
    from app.models.usuario import Usuario
    from app.models.promocion import Promocion
    from app.models.cliente import Cliente
    from app.services.actualizaciones_datos import registrar_actualizacion_datos
    from sqlalchemy.future import select
    from sqlalchemy import delete
    from decimal import Decimal

    async with SessionLocal() as db:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
            
            def safe_float(v):
                try:
                    return float(v) if v is not None else 0.0
                except (ValueError, TypeError):
                    return 0.0

            def safe_date(v):
                if hasattr(v, 'date'):
                    return v.date()
                if isinstance(v, str):
                    try:
                        return datetime.strptime(v.strip().split(' ')[0], '%Y-%m-%d').date()
                    except Exception:
                        pass
                return v

            users_res = await db.execute(select(Usuario))
            users = users_res.scalars().all()
            users_by_code = {
                _normalize_seller_text(user.codigo_vendedor): user.id
                for user in users
                if _normalize_seller_text(user.codigo_vendedor)
            }
            users_by_name = _seller_ids_by_name(users)

            existing_quotes = (await db.execute(select(Cotizacion))).scalars().all()
            existing_by_number: dict[str, Cotizacion] = {}
            for existing in existing_quotes:
                number = _excel_identifier(existing.numero_cotizacion)
                if number:
                    existing_by_number[number] = existing

            retained_ids: set[UUID] = set()

            multi_sheets = _detect_multi_sheet_worksheets(wb)
            if multi_sheets is not None:
                # -------------------------------------------------------------
                # NUEVO FORMATO MULTI-HOJA DINÁMICO (Cabecera + Detalle)
                # -------------------------------------------------------------
                ws_ventas, v_cols, ws_cot, c_cols = multi_sheets

                # Cargar catálogo de promociones
                promos_res = await db.execute(select(Promocion))
                promos = promos_res.scalars().all()
                promos_by_code = {
                    normalize_text(p.codigo_material): p
                    for p in promos
                    if p.codigo_material and normalize_text(p.codigo_material)
                }

                # Cargar clientes existentes para enriquecer datos de contacto
                clients_res = await db.execute(select(Cliente))
                clients = clients_res.scalars().all()
                clients_by_num = {c.numero_cliente.strip(): c for c in clients if c.numero_cliente and c.numero_cliente.strip()}
                clients_by_nom = {c.nombre.strip(): c for c in clients if c.nombre and c.nombre.strip()}

                # Índices Hoja Detalle (ws_cot)
                cot_idx_fec = c_cols.get("fecha_registro")
                cot_idx_org = c_cols.get("organizacion_ventas")
                cot_idx_num = c_cols.get("numero_cotizacion")
                cot_idx_abc = c_cols.get("indicador_abcf")
                cot_idx_sku = c_cols.get("codigo_material")
                cot_idx_des = c_cols.get("descripcion")
                cot_idx_udm = c_cols.get("unidad_medida")
                cot_idx_prc = c_cols.get("precio")

                cot_idx_cli_nom = c_cols.get("cliente_nombre")
                cot_idx_cli_num = c_cols.get("numero_cliente")
                cot_idx_vend_nom = c_cols.get("vendedor_nombre")
                cot_idx_vend_cod = c_cols.get("vendedor_codigo")
                cot_idx_canal = c_cols.get("canal")
                cot_idx_tel = c_cols.get("telefono")
                cot_idx_cel = c_cols.get("celular")
                cot_idx_email = c_cols.get("email")

                # Agrupar cotizaciones y partidas desde la hoja detalle
                quotes_data: dict[str, dict[str, Any]] = {}
                for row in ws_cot.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    num_cot = _excel_identifier(row[cot_idx_num]) if cot_idx_num is not None else None
                    if not num_cot:
                        continue

                    sku = _excel_identifier(row[cot_idx_sku]) if cot_idx_sku is not None else None
                    if not sku:
                        continue

                    precio = safe_float(row[cot_idx_prc]) if cot_idx_prc is not None else 0.0
                    fecha_reg = safe_date(row[cot_idx_fec]) if cot_idx_fec is not None else None
                    org_v = _excel_identifier(row[cot_idx_org]) if cot_idx_org is not None else None
                    abc = _excel_identifier(row[cot_idx_abc]) if cot_idx_abc is not None else None
                    desc = _excel_identifier(row[cot_idx_des]) if cot_idx_des is not None else None
                    udm = _excel_identifier(row[cot_idx_udm]) if cot_idx_udm is not None else None

                    if num_cot not in quotes_data:
                        quotes_data[num_cot] = {
                            "numero_cotizacion": num_cot,
                            "fecha_registro": fecha_reg,
                            "organizacion_ventas": org_v,
                            "total": Decimal("0.00"),
                            "items_data": [],
                        }

                    # Extraer metadatos opcionales si vienen en la hoja detalle
                    if cot_idx_cli_nom is not None and not quotes_data[num_cot].get("cliente_nombre"):
                        c_cli = _excel_identifier(row[cot_idx_cli_nom])
                        if c_cli:
                            quotes_data[num_cot]["cliente_nombre"] = c_cli
                    if cot_idx_cli_num is not None and not quotes_data[num_cot].get("numero_cliente"):
                        c_num = _excel_identifier(row[cot_idx_cli_num])
                        if c_num:
                            quotes_data[num_cot]["numero_cliente"] = c_num
                    if cot_idx_vend_nom is not None and not quotes_data[num_cot].get("vendedor_nombre"):
                        c_vnom = _excel_identifier(row[cot_idx_vend_nom])
                        if c_vnom:
                            quotes_data[num_cot]["vendedor_nombre"] = c_vnom
                    if cot_idx_vend_cod is not None and not quotes_data[num_cot].get("vendedor_codigo"):
                        c_vcod = _excel_identifier(row[cot_idx_vend_cod])
                        if c_vcod:
                            quotes_data[num_cot]["vendedor_codigo"] = c_vcod
                    if cot_idx_canal is not None and not quotes_data[num_cot].get("canal"):
                        c_can = _excel_identifier(row[cot_idx_canal])
                        if c_can:
                            quotes_data[num_cot]["canal"] = c_can

                    if cot_idx_tel is not None or cot_idx_cel is not None or cot_idx_email is not None:
                        tel_val = _excel_identifier(row[cot_idx_tel]) if cot_idx_tel is not None else None
                        cel_val = _excel_identifier(row[cot_idx_cel]) if cot_idx_cel is not None else None
                        em_val = _excel_identifier(row[cot_idx_email]) if cot_idx_email is not None else None
                        if "contact_data" not in quotes_data[num_cot]:
                            quotes_data[num_cot]["contact_data"] = {}
                        if tel_val and not quotes_data[num_cot]["contact_data"].get("telefono"):
                            quotes_data[num_cot]["contact_data"]["telefono"] = tel_val
                        if cel_val and not quotes_data[num_cot]["contact_data"].get("celular"):
                            quotes_data[num_cot]["contact_data"]["celular"] = cel_val
                        if em_val and not quotes_data[num_cot]["contact_data"].get("email"):
                            quotes_data[num_cot]["contact_data"]["email"] = em_val

                    sku_norm = normalize_text(sku)
                    promo_match = promos_by_code.get(sku_norm)
                    es_promo = promo_match is not None
                    prc_promo = Decimal(str(promo_match.precio_promocion)) if promo_match and promo_match.precio_promocion else None

                    dec_price = Decimal(str(round(precio, 2)))
                    quotes_data[num_cot]["total"] += dec_price
                    quotes_data[num_cot]["items_data"].append({
                        "codigo_material": sku,
                        "descripcion": desc,
                        "indicador_abcf": abc,
                        "unidad_medida": udm,
                        "precio_venta": dec_price,
                        "cantidad_cotizada": Decimal("1.000"),
                        "importe_cotizado": dec_price,
                        "es_promocion": es_promo,
                        "precio_promocion": prc_promo,
                    })

                # Índices Hoja Cabecera / Ventas (ws_ventas)
                v_idx_cot_fol = v_cols.get("numero_cotizacion")
                v_idx_fac_fol = v_cols.get("numero_factura")
                v_idx_fec_fac = v_cols.get("fecha_factura")
                v_idx_fec_reg = v_cols.get("fecha_registro")
                v_idx_cli_num = v_cols.get("numero_cliente")
                v_idx_cli_nom = v_cols.get("cliente_nombre")
                v_idx_plazo = v_cols.get("plazo_entrega")
                v_idx_hora = v_cols.get("hora_facturacion")
                v_idx_margen = v_cols.get("margen")
                v_idx_grp_vend = v_cols.get("grupo_vendedores")
                v_idx_vend_nom = v_cols.get("vendedor_nombre")
                v_idx_vend_cod = v_cols.get("vendedor_codigo")
                v_idx_canal = v_cols.get("canal")
                v_idx_org = v_cols.get("organizacion_ventas")
                v_idx_tot = v_cols.get("precio")
                v_idx_sku = v_cols.get("codigo_material")
                v_idx_des = v_cols.get("descripcion")
                v_idx_cant_fac = v_cols.get("cantidad_facturada")
                v_idx_imp_fac = v_cols.get("importe_facturado")
                v_idx_abcf = v_cols.get("indicador_abcf")
                v_idx_tel = v_cols.get("telefono")
                v_idx_cel = v_cols.get("celular")
                v_idx_email = v_cols.get("email")

                ventas_by_cot: dict[str, dict[str, Any]] = {}
                for row in ws_ventas.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    fac_fol = _excel_identifier(row[v_idx_fac_fol]) if v_idx_fac_fol is not None else None
                    cot_fol = _excel_identifier(row[v_idx_cot_fol]) if v_idx_cot_fol is not None else None
                    if not fac_fol and not cot_fol:
                        continue

                    target_folio = cot_fol or fac_fol

                    fec_fac = safe_date(row[v_idx_fec_fac]) if v_idx_fec_fac is not None else None
                    fec_reg = safe_date(row[v_idx_fec_reg]) if v_idx_fec_reg is not None else None
                    cli_num = _excel_identifier(row[v_idx_cli_num]) if v_idx_cli_num is not None else None
                    plazo = _excel_identifier(row[v_idx_plazo]) if v_idx_plazo is not None else None
                    cli_nom = _excel_identifier(row[v_idx_cli_nom]) if v_idx_cli_nom is not None else None

                    hora_val = row[v_idx_hora] if v_idx_hora is not None else None
                    hora_str = str(hora_val) if hora_val is not None else None
                    margen_val = safe_float(row[v_idx_margen]) if v_idx_margen is not None else None
                    grp_vend = _excel_identifier(row[v_idx_grp_vend]) if v_idx_grp_vend is not None else None
                    vend_nom = _excel_identifier(row[v_idx_vend_nom]) if v_idx_vend_nom is not None else None
                    vend_cod = _excel_identifier(row[v_idx_vend_cod]) if v_idx_vend_cod is not None else None
                    canal_val = _excel_identifier(row[v_idx_canal]) if v_idx_canal is not None else None
                    org_val = _excel_identifier(row[v_idx_org]) if v_idx_org is not None else None

                    # Datos de la partida facturada (si la hoja Ventas trae desglose)
                    v_sku = _excel_identifier(row[v_idx_sku]) if v_idx_sku is not None else None
                    v_desc = _excel_identifier(row[v_idx_des]) if v_idx_des is not None else None
                    v_abcf = _excel_identifier(row[v_idx_abcf]) if v_idx_abcf is not None else None
                    v_cant_raw = safe_float(row[v_idx_cant_fac]) if v_idx_cant_fac is not None else 1.0
                    v_imp_raw = (
                        safe_float(row[v_idx_imp_fac]) if v_idx_imp_fac is not None
                        else (safe_float(row[v_idx_tot]) if v_idx_tot is not None else 0.0)
                    )

                    dec_cant_fac = Decimal(str(round(v_cant_raw, 3))) if v_cant_raw is not None else Decimal("1.000")
                    dec_imp_fac = Decimal(str(round(v_imp_raw, 2))) if v_imp_raw is not None else Decimal("0.00")

                    if target_folio not in ventas_by_cot:
                        v_info = {
                            "fecha_factura": fec_fac,
                            "fecha_registro": fec_reg,
                            "numero_cliente": cli_num,
                            "cliente_nombre": cli_nom,
                            "plazo_entrega": plazo,
                            "numero_factura": fac_fol,
                            "hora_facturacion": hora_str,
                            "margen": Decimal(str(round(margen_val, 3))) if margen_val is not None else None,
                            "margen_items": [(margen_val, float(dec_imp_fac))] if margen_val is not None else [],
                            "grupo_vendedores": grp_vend,
                            "vendedor_nombre": vend_nom,
                            "vendedor_codigo": vend_cod,
                            "canal": canal_val,
                            "organizacion_ventas": org_val,
                            "total_facturado": Decimal("0.00"),
                            "items_facturados": [],
                        }

                        if v_idx_tel is not None or v_idx_cel is not None or v_idx_email is not None:
                            tel_v = _excel_identifier(row[v_idx_tel]) if v_idx_tel is not None else None
                            cel_v = _excel_identifier(row[v_idx_cel]) if v_idx_cel is not None else None
                            em_v = _excel_identifier(row[v_idx_email]) if v_idx_email is not None else None
                            v_info["contact_data"] = {
                                "telefono": tel_v,
                                "celular": cel_v,
                                "email": em_v,
                            }

                        ventas_by_cot[target_folio] = v_info
                    else:
                        v_info = ventas_by_cot[target_folio]
                        if not v_info.get("cliente_nombre") and cli_nom:
                            v_info["cliente_nombre"] = cli_nom
                        if not v_info.get("numero_cliente") and cli_num:
                            v_info["numero_cliente"] = cli_num
                        if not v_info.get("vendedor_nombre") and vend_nom:
                            v_info["vendedor_nombre"] = vend_nom
                        if not v_info.get("vendedor_codigo") and vend_cod:
                            v_info["vendedor_codigo"] = vend_cod
                        if not v_info.get("numero_factura") and fac_fol:
                            v_info["numero_factura"] = fac_fol
                        if not v_info.get("fecha_factura") and fec_fac:
                            v_info["fecha_factura"] = fec_fac
                        if margen_val is not None:
                            v_info["margen_items"].append((margen_val, float(dec_imp_fac)))

                    v_info["total_facturado"] += dec_imp_fac
                    if v_sku:
                        v_info["items_facturados"].append({
                            "codigo_material": v_sku,
                            "descripcion": v_desc,
                            "cantidad_facturada": dec_cant_fac,
                            "importe_facturado": dec_imp_fac,
                            "indicador_abcf": v_abcf,
                            "margen": margen_val,
                        })

                # Calcular margen ponderado
                for v_item in ventas_by_cot.values():
                    m_items = v_item.pop("margen_items", [])
                    if m_items:
                        total_weight = sum(max(w, 1.0) for _, w in m_items)
                        weighted_margin = sum(m * max(w, 1.0) for m, w in m_items) / total_weight
                        v_item["margen"] = Decimal(str(round(weighted_margin, 3)))

                BATCH_QUOTES = 2000
                quote_keys = list(dict.fromkeys(list(quotes_data.keys()) + list(ventas_by_cot.keys())))

                for q_start in range(0, len(quote_keys), BATCH_QUOTES):
                    batch_keys = quote_keys[q_start : q_start + BATCH_QUOTES]
                    batch_quote_ids: list[UUID] = []
                    batch_items: list[dict[str, Any]] = []

                    for num_cot in batch_keys:
                        q_data = quotes_data.get(num_cot) or {}
                        v_info = ventas_by_cot.get(num_cot) or {}

                        cli_nom = v_info.get("cliente_nombre") or q_data.get("cliente_nombre")
                        cli_num = v_info.get("numero_cliente") or q_data.get("numero_cliente")
                        vend_nom = v_info.get("vendedor_nombre") or q_data.get("vendedor_nombre")
                        vend_cod = v_info.get("vendedor_codigo") or q_data.get("vendedor_codigo")
                        canal_val = v_info.get("canal") or q_data.get("canal")
                        org_val = q_data.get("organizacion_ventas") or v_info.get("organizacion_ventas")
                        fecha_reg = q_data.get("fecha_registro") or v_info.get("fecha_registro")

                        existing = existing_by_number.get(num_cot)
                        if existing:
                            if not cli_nom or cli_nom in ("Cliente Desconocido", "Cliente sin registrar"):
                                if existing.cliente_nombre and existing.cliente_nombre not in ("Cliente Desconocido", "Cliente sin registrar"):
                                    cli_nom = existing.cliente_nombre
                            if not cli_num and existing.numero_cliente:
                                cli_num = existing.numero_cliente
                            if not vend_nom and existing.vendedor_nombre:
                                vend_nom = existing.vendedor_nombre
                            if not vend_cod and existing.grupo_vendedores:
                                vend_cod = existing.grupo_vendedores

                        vend_id = (
                            (users_by_code.get(_normalize_seller_text(vend_cod)) if vend_cod else None)
                            or (users_by_name.get(_normalize_seller_text(vend_nom)) if vend_nom else None)
                            or (users_by_name.get(_normalize_seller_text(vend_cod)) if vend_cod else None)
                            or (users_by_code.get(_normalize_seller_text(vend_nom)) if vend_nom else None)
                            or (existing.vendedor_id if existing and existing.vendedor_id else None)
                        )

                        if vend_id and not vend_nom:
                            u_match = next((u for u in users if u.id == vend_id), None)
                            if u_match and u_match.nombre_completo:
                                vend_nom = u_match.nombre_completo

                        cli_obj = (clients_by_num.get(cli_num) if cli_num else None) or (clients_by_nom.get(cli_nom) if cli_nom else None)
                        if cli_obj:
                            if not cli_nom or cli_nom in ("Cliente Desconocido", "Cliente sin registrar"):
                                cli_nom = cli_obj.nombre
                            if not cli_num:
                                cli_num = cli_obj.numero_cliente

                        if (not cli_nom or cli_nom in ("Cliente Desconocido", "Cliente sin registrar")) and cli_num:
                            cli_nom = f"Cliente #{cli_num}"

                        contact_data = dict(v_info.get("contact_data") or q_data.get("contact_data") or {})
                        if cli_obj:
                            if cli_obj.email and not contact_data.get("email"):
                                contact_data["email"] = cli_obj.email
                            if cli_obj.telefono and not contact_data.get("telefono"):
                                contact_data["telefono"] = cli_obj.telefono
                            if cli_obj.celular and not contact_data.get("celular"):
                                contact_data["celular"] = cli_obj.celular
                            if cli_obj.nombre_contacto and not contact_data.get("nombre_contacto"):
                                contact_data["nombre_contacto"] = cli_obj.nombre_contacto

                        # Cruzar partidas cotizadas (q_data) con partidas facturadas (v_info)
                        items_cot = q_data.get("items_data", [])
                        items_fac = v_info.get("items_facturados", [])

                        items_fac_by_sku: dict[str, list[dict[str, Any]]] = defaultdict(list)
                        for f_it in items_fac:
                            items_fac_by_sku[normalize_text(f_it["codigo_material"])].append(f_it)

                        final_quote_items: list[dict[str, Any]] = []
                        matched_fac_skus: set[str] = set()

                        # 1. Partidas cotizadas
                        for it_data in items_cot:
                            sku_norm = normalize_text(it_data["codigo_material"])
                            fac_entries = items_fac_by_sku.get(sku_norm)

                            if fac_entries:
                                matched_fac_skus.add(sku_norm)
                                it_fac_qty = sum((e["cantidad_facturada"] for e in fac_entries), Decimal("0.000"))
                                it_fac_amt = sum((e["importe_facturado"] for e in fac_entries), Decimal("0.00"))
                            elif v_info.get("numero_factura") and not items_fac:
                                # Archivo legado donde Ventas no traía detalle de SKU
                                it_fac_qty = it_data["cantidad_cotizada"]
                                it_fac_amt = it_data["precio_venta"]
                            else:
                                it_fac_qty = Decimal("0.000")
                                it_fac_amt = Decimal("0.00")

                            final_quote_items.append({
                                "codigo_material": it_data["codigo_material"],
                                "descripcion": it_data["descripcion"],
                                "indicador_abcf": it_data["indicador_abcf"],
                                "unidad_medida": it_data["unidad_medida"],
                                "precio_venta": it_data["precio_venta"],
                                "cantidad_cotizada": it_data["cantidad_cotizada"],
                                "importe_cotizado": it_data["importe_cotizado"],
                                "cantidad_facturada": it_fac_qty,
                                "importe_facturado": it_fac_amt,
                                "es_promocion": it_data["es_promocion"],
                                "precio_promocion": it_data["precio_promocion"],
                            })

                        # 2. Partidas facturadas adicionales (venta cruzada)
                        for f_it in items_fac:
                            sku_norm = normalize_text(f_it["codigo_material"])
                            if sku_norm not in matched_fac_skus:
                                matched_fac_skus.add(sku_norm)
                                promo_match = promos_by_code.get(sku_norm)
                                unit_prc = (f_it["importe_facturado"] / f_it["cantidad_facturada"]) if f_it["cantidad_facturada"] > 0 else f_it["importe_facturado"]
                                final_quote_items.append({
                                    "codigo_material": f_it["codigo_material"],
                                    "descripcion": f_it["descripcion"],
                                    "indicador_abcf": f_it["indicador_abcf"],
                                    "unidad_medida": "PZA",
                                    "precio_venta": Decimal(str(round(unit_prc, 2))),
                                    "cantidad_cotizada": Decimal("0.000"),
                                    "importe_cotizado": Decimal("0.00"),
                                    "cantidad_facturada": f_it["cantidad_facturada"],
                                    "importe_facturado": f_it["importe_facturado"],
                                    "es_promocion": promo_match is not None,
                                    "precio_promocion": Decimal(str(promo_match.precio_promocion)) if promo_match and promo_match.precio_promocion else None,
                                })

                        # Métricas agregadas
                        cotizados_skus = [it["codigo_material"] for it in final_quote_items if it["cantidad_cotizada"] > 0]
                        facturados_skus = [it["codigo_material"] for it in final_quote_items if it["cantidad_facturada"] > 0]

                        mat_cot_str = ", ".join(dict.fromkeys(cotizados_skus)) if cotizados_skus else None
                        mat_fac_str = ", ".join(dict.fromkeys(facturados_skus)) if facturados_skus else None

                        sum_cotizado = sum((it["importe_cotizado"] for it in final_quote_items), Decimal("0.00"))
                        sum_facturado = sum((it["importe_facturado"] for it in final_quote_items), Decimal("0.00"))

                        total_val = q_data.get("total") if (q_data.get("total") is not None and q_data.get("total") > 0) else (sum_cotizado if sum_cotizado > 0 else (v_info.get("total_facturado") or Decimal("0.00")))

                        if v_info.get("numero_factura"):
                            if sum_facturado > 0:
                                importe_fac = sum_facturado
                            elif v_info.get("total_facturado") and v_info.get("total_facturado") > 0:
                                importe_fac = v_info["total_facturado"]
                            else:
                                importe_fac = total_val
                        else:
                            importe_fac = None

                        pct_mat = None
                        if v_info.get("numero_factura") or facturados_skus:
                            if cotizados_skus:
                                intersect = set(facturados_skus) & set(cotizados_skus)
                                pct_mat = Decimal(str(round((len(intersect) / len(set(cotizados_skus))) * 100, 2)))
                            elif facturados_skus:
                                pct_mat = Decimal("100.00")

                        pct_imp = None
                        if importe_fac is not None and total_val > 0:
                            pct_imp = Decimal(str(round((float(importe_fac) / float(total_val)) * 100, 2)))

                        items_json = [
                            {
                                "producto": it["codigo_material"],
                                "codigo_material": it["codigo_material"],
                                "descripcion": it.get("descripcion"),
                                "cantidad": float(it["cantidad_cotizada"]) if it["cantidad_cotizada"] > 0 else float(it["cantidad_facturada"]),
                                "precio_unitario": float(it["precio_venta"]),
                                "indicador_abcf": it.get("indicador_abcf"),
                                "unidad_medida": it.get("unidad_medida"),
                                "es_promocion": it.get("es_promocion", False),
                                "precio_promocion": float(it["precio_promocion"]) if it.get("precio_promocion") else None,
                                "cantidad_facturada": float(it["cantidad_facturada"]),
                                "importe_facturado": float(it["importe_facturado"]),
                            }
                            for it in final_quote_items
                        ]

                        quote_fields = {
                            "numero_cotizacion": num_cot,
                            "fecha_registro": fecha_reg,
                            "organizacion_ventas": org_val,
                            "vendedor_id": vend_id,
                            "vendedor_nombre": vend_nom,
                            "numero_cliente": cli_num,
                            "cliente_nombre": cli_nom or "Cliente Desconocido",
                            "datos_contacto": contact_data,
                            "items": items_json,
                            "materiales_cotizados": mat_cot_str,
                            "materiales_facturados": mat_fac_str,
                            "porcentaje_materiales": pct_mat,
                            "porcentaje_importe": pct_imp,
                            "total": total_val,
                            "numero_factura": v_info.get("numero_factura"),
                            "fecha_factura": v_info.get("fecha_factura"),
                            "hora_facturacion": v_info.get("hora_facturacion"),
                            "margen": v_info.get("margen"),
                            "grupo_vendedores": v_info.get("grupo_vendedores") or vend_cod,
                            "plazo_entrega": v_info.get("plazo_entrega"),
                            "canal": canal_val,
                            "importe_facturado": importe_fac,
                        }

                        existing = existing_by_number.get(num_cot)
                        if existing:
                            final_fields = _merge_quote_fields(existing, quote_fields)
                            _apply_imported_quote_values(existing, final_fields)
                            target_quote_id = existing.id
                            retained_ids.add(existing.id)
                        else:
                            target_quote_id = uuid4()
                            new_q = Cotizacion(id=target_quote_id, **quote_fields)
                            db.add(new_q)
                            existing_by_number[num_cot] = new_q
                            retained_ids.add(target_quote_id)

                        batch_quote_ids.append(target_quote_id)

                        for it_data in final_quote_items:
                            batch_items.append({
                                "id": uuid4(),
                                "cotizacion_id": target_quote_id,
                                "codigo_material": it_data["codigo_material"],
                                "descripcion": it_data["descripcion"],
                                "indicador_abcf": it_data["indicador_abcf"],
                                "unidad_medida": it_data["unidad_medida"],
                                "precio_venta": it_data["precio_venta"],
                                "cantidad_cotizada": it_data["cantidad_cotizada"],
                                "importe_cotizado": it_data["importe_cotizado"],
                                "cantidad_facturada": it_data["cantidad_facturada"],
                                "importe_facturado": it_data["importe_facturado"],
                                "es_promocion": it_data["es_promocion"],
                                "precio_promocion": it_data["precio_promocion"],
                            })

                    # Flush cotizaciones del lote
                    await db.flush()

                    # Eliminar partidas anteriores de este lote
                    if batch_quote_ids:
                        for i in range(0, len(batch_quote_ids), 1000):
                            chunk_del = batch_quote_ids[i : i + 1000]
                            await db.execute(delete(CotizacionItem).where(CotizacionItem.cotizacion_id.in_(chunk_del)))

                    # Insertar partidas de este lote
                    if batch_items:
                        for i in range(0, len(batch_items), 5000):
                            chunk_ins = batch_items[i : i + 5000]
                            await db.execute(insert(CotizacionItem), chunk_ins)

                    # Commit incremental para rotar logs WAL y liberar espacio de disco
                    await db.commit()

            else:
                # -------------------------------------------------------------
                # FORMATO LEGADO (HOJA ÚNICA / MÚLTIPLES MESES)
                # -------------------------------------------------------------
                worksheets_and_indices = _find_all_quote_worksheets(wb)
                seen_numbers: set[str] = set()
                new_quotes = []

                for ws, column_indices in worksheets_and_indices:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row:
                            continue
                        num_cot_val = _excel_identifier(row[column_indices["numero_cotizacion"]])
                        if not num_cot_val:
                            continue

                        if num_cot_val in seen_numbers:
                            continue
                        seen_numbers.add(num_cot_val)

                    fecha_reg = safe_date(row[column_indices["fecha_registro"]])
                    org_ventas = _excel_identifier(row[column_indices["organizacion_ventas"]])
                    canal_val = _excel_identifier(row[column_indices["canal"]])
                    vend_codigo = _excel_identifier(row[column_indices["vendedor_codigo"]])
                    vend_nombre = _excel_identifier(row[column_indices["vendedor_nombre"]])
                    num_cliente = _excel_identifier(row[column_indices["numero_cliente"]])
                    cliente_nombre = _excel_identifier(row[column_indices["cliente_nombre"]])
                    telefono = _excel_identifier(row[column_indices["telefono"]])
                    celular = _excel_identifier(row[column_indices["celular"]])
                    email = _excel_identifier(row[column_indices["email"]])
                    num_factura = _excel_identifier(row[column_indices["numero_factura"]])
                    fecha_fac = safe_date(row[column_indices["fecha_factura"]])

                    importe_cot = safe_float(row[column_indices["importe_cotizado"]])
                    importe_fac = safe_float(row[column_indices["importe_facturado"]])
                    pct_importe = safe_float(row[column_indices["porcentaje_importe"]])
                    mat_cot = _excel_identifier(row[column_indices["materiales_cotizados"]])
                    mat_fac = _excel_identifier(row[column_indices["materiales_facturados"]])
                    pct_mat = safe_float(row[column_indices["porcentaje_materiales"]])

                    vendedor_id = (
                        users_by_code.get(_normalize_seller_text(vend_codigo))
                        or users_by_name.get(_normalize_seller_text(vend_nombre))
                    )

                    imported_values = {
                        "numero_cotizacion": num_cot_val,
                        "fecha_registro": fecha_reg,
                        "organizacion_ventas": org_ventas,
                        "canal": canal_val,
                        "vendedor_id": vendedor_id,
                        "vendedor_nombre": vend_nombre,
                        "numero_cliente": num_cliente,
                        "cliente_nombre": cliente_nombre or "Cliente Desconocido",
                        "datos_contacto": {"email": email, "telefono": telefono, "celular": celular},
                        "items": [],
                        "numero_factura": num_factura,
                        "fecha_factura": fecha_fac,
                        "total": Decimal(str(round(importe_cot, 2))),
                        "importe_facturado": Decimal(str(round(importe_fac, 2))) if importe_fac else None,
                        "porcentaje_importe": Decimal(str(round(pct_importe, 2))) if pct_importe else None,
                        "materiales_cotizados": mat_cot,
                        "materiales_facturados": mat_fac,
                        "porcentaje_materiales": Decimal(str(round(pct_mat, 2))) if pct_mat else None,
                    }
                    existing_quote = existing_by_number.get(num_cot_val)
                    if existing_quote is not None:
                        merged_values = _merge_quote_fields(existing_quote, imported_values)
                        _apply_imported_quote_values(existing_quote, merged_values)
                        retained_ids.add(existing_quote.id)
                    else:
                        new_id = uuid4()
                        new_quotes.append(Cotizacion(id=new_id, **imported_values))
                        retained_ids.add(new_id)

                BATCH_SIZE = 1000
                for i in range(0, len(new_quotes), BATCH_SIZE):
                    db.add_all(new_quotes[i : i + BATCH_SIZE])
                    await db.flush()

            # Limpieza de registros obsoletos
            stale_ids = _stale_imported_quote_ids(existing_quotes, retained_ids)
            if stale_ids:
                stale_list = list(stale_ids)
                for i in range(0, len(stale_list), 1000):
                    chunk = stale_list[i : i + 1000]
                    await db.execute(delete(Cotizacion).where(Cotizacion.id.in_(chunk)))

            await registrar_actualizacion_datos(db, "cotizaciones", uploaded_by_id)
            await db.commit()
            print("Background upload finished successfully.")
            return None

        except Exception as e:
            await db.rollback()
            err = f"Error procesando cotizaciones: {str(e)}"
            print(err)
            return err

