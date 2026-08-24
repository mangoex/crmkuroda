from collections import defaultdict
from datetime import datetime
import io
import re

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy import delete, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select

from app.core.database import get_db
from app.core.security import RoleChecker, get_current_user
from app.models.sobrepedido import Sobrepedido
from app.models.por_entregar import PorEntregar
from app.models.usuario import Usuario
from app.models.cliente import Cliente
from app.services.sobrepedidos_classifier import classify_por_entregar, classify_sobrepedido, clean_text
from app.services.jerarquia import get_codigos_vendedores_visibles

router = APIRouter()

require_admin_or_gerente = RoleChecker(["admin", "gerente", "logistica"])


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    text = clean_text(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text


def safe_float(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def safe_int(value):
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def normalize_key(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value)


def find_col_index(headers, patterns):
    normalized_headers = [clean_text(h) for h in headers]
    for pattern in patterns:
        for idx, header in enumerate(normalized_headers):
            if header and re.search(pattern, header, re.IGNORECASE):
                return idx
    raise ValueError(f"No se encontro la columna requerida para {patterns}")


def get_cell(row, index, default=None):
    if index >= len(row):
        return default
    return row[index]


@router.get("/")
async def list_sobrepedidos(
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    query = select(Sobrepedido)
    if current_user.rol == "vendedor":
        codigos_visibles = await get_codigos_vendedores_visibles(db, current_user)
        if codigos_visibles:
            query = query.filter(Sobrepedido.vendedor_codigo.in_(codigos_visibles))
        else:
            # Vendedor sin codigo: no ve nada (no hay forma de matchear)
            query = query.filter(Sobrepedido.vendedor_codigo == "__NO_MATCH__")

    result = await db.execute(query)
    records = result.scalars().all()

    # Enriquecer con datos de contacto del catálogo de Clientes
    client_numbers = {r.numero_cliente.strip() for r in records if r.numero_cliente and r.numero_cliente.strip()}
    client_names = {r.cliente_nombre.strip() for r in records if r.cliente_nombre and r.cliente_nombre.strip()}
    clients_by_number = {}
    clients_by_name = {}
    if client_numbers or client_names:
        conditions = []
        if client_numbers:
            conditions.append(Cliente.numero_cliente.in_(client_numbers))
        if client_names:
            conditions.append(Cliente.nombre.in_(client_names))
        catalog_clients = (await db.execute(select(Cliente).where(or_(*conditions)))).scalars().all()
        for cl in catalog_clients:
            if cl.numero_cliente and cl.numero_cliente.strip():
                clients_by_number[cl.numero_cliente.strip()] = cl
            if cl.nombre and cl.nombre.strip():
                clients_by_name[cl.nombre.strip()] = cl

    data = []
    for r in records:
        d = r.to_dict()
        cl = clients_by_number.get(r.numero_cliente.strip() if r.numero_cliente else "") or clients_by_name.get(r.cliente_nombre.strip() if r.cliente_nombre else "")
        tel = (cl.telefono or "").strip() if cl and cl.telefono else None
        cel = (cl.celular or "").strip() if cl and cl.celular else None
        d["contacto"] = {
            "telefono": tel,
            "celular": cel,
            "contacto_preferente": cel or tel,
            "email": (cl.email or "").strip() if cl and cl.email else None,
            "nombre_contacto": (cl.nombre_contacto or "").strip() if cl and cl.nombre_contacto else None,
        }
        data.append(d)

    return {"status": "success", "data": data}


@router.post("/upload", status_code=status.HTTP_201_CREATED)
async def upload_sobrepedidos(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(require_admin_or_gerente),
):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")

    contents = await file.read()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        required_sheets = {"VA05", "VL06O"}
        missing = required_sheets.difference(wb.sheetnames)
        if missing:
            missing_text = ", ".join(sorted(missing))
            raise HTTPException(
                status_code=400,
                detail=f"El archivo debe contener las hojas VA05 y VL06O. Faltan: {missing_text}",
            )

        users_res = await db.execute(select(Usuario))
        users_by_code = {
            clean_text(user.codigo_vendedor): user
            for user in users_res.scalars().all()
            if clean_text(user.codigo_vendedor)
        }

        rows_added = await _replace_sobrepedidos_from_va05(db, wb["VA05"], users_by_code)
        delivery_rows_added = await _replace_por_entregar_from_vl06o(db, wb["VL06O"], users_by_code)

        await db.commit()
        return {
            "status": "success",
            "message": f"Se cargaron {rows_added} lineas de sobrepedidos y {delivery_rows_added} lineas por entregar.",
        }
    except HTTPException:
        await db.rollback()
        raise
    except Exception as exc:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error procesando el archivo: {str(exc)}")


def _parse_vl06o(ws):
    headers = [cell.value for cell in ws[1]]
    idx_factura = find_col_index(headers, [r"^factura$"])
    idx_codigo = find_col_index(headers, [r"^codigo$", r"c[oó]digo"])
    idx_cantidad = find_col_index(headers, [r"cantidad.*entregar"])
    idx_fecha = find_col_index(headers, [r"fecha.*disponibilidad"])
    idx_dias = find_col_index(headers, [r"dias.*disponible", r"d[ií]as.*disponible"])

    by_factura = defaultdict(list)
    by_factura_codigo = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        factura = normalize_key(get_cell(row, idx_factura))
        codigo = normalize_key(get_cell(row, idx_codigo))
        if not factura:
            continue

        item = {
            "factura": factura,
            "codigo": codigo,
            "cantidad": safe_float(get_cell(row, idx_cantidad)),
            "fecha_disponibilidad": parse_date(get_cell(row, idx_fecha)),
            "dias_disponible": safe_int(get_cell(row, idx_dias)),
        }
        by_factura[factura].append(item)
        if codigo:
            by_factura_codigo[(factura, codigo)].append(item)

    return {"by_factura": by_factura, "by_factura_codigo": by_factura_codigo}


async def _replace_sobrepedidos_from_va05(db, ws, users_by_code):
    headers = [cell.value for cell in ws[1]]
    idx_fecha_venta = find_col_index(headers, [r"fecha.*venta"])
    idx_factura = find_col_index(headers, [r"^factura$"])
    idx_vendedor = find_col_index(headers, [r"^vendedor$"])
    idx_num_cliente = find_col_index(headers, [r"num.*cliente"])
    idx_cliente = find_col_index(headers, [r"nombre.*cliente"])
    idx_proveedor = find_col_index(headers, [r"proveedor"])
    idx_codigo = find_col_index(headers, [r"^codigo$", r"c[oó]digo"])
    idx_indicador = find_col_index(headers, [r"indicador"])
    idx_producto = find_col_index(headers, [r"producto"])
    idx_grupo = find_col_index(headers, [r"grupo"])
    idx_pendiente = find_col_index(headers, [r"cantidad.*pendiente"])
    idx_estatus = find_col_index(headers, [r"estatus.*compras"])

    await db.execute(delete(Sobrepedido))

    rows_added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        factura = normalize_key(get_cell(row, idx_factura))
        codigo = normalize_key(get_cell(row, idx_codigo))
        cantidad_pendiente = safe_float(get_cell(row, idx_pendiente))

        if not factura or not codigo or cantidad_pendiente <= 0:
            continue

        status_result = classify_sobrepedido(
            estatus_compras=get_cell(row, idx_estatus),
            cantidad_pendiente=cantidad_pendiente,
        )

        vendedor_codigo = normalize_key(get_cell(row, idx_vendedor))
        vendedor = users_by_code.get(vendedor_codigo)
        vendedor_nombre = vendedor.nombre_completo if vendedor else vendedor_codigo

        record = Sobrepedido(
            id_pedido_erp=safe_int(factura),
            factura=factura,
            fecha_venta=parse_date(get_cell(row, idx_fecha_venta)),
            fecha_pedido=parse_date(get_cell(row, idx_fecha_venta)),
            vendedor_codigo=vendedor_codigo,
            vendedor_nombre=vendedor_nombre,
            numero_cliente=normalize_key(get_cell(row, idx_num_cliente)),
            cliente_nombre=clean_text(get_cell(row, idx_cliente)),
            proveedor=clean_text(get_cell(row, idx_proveedor)),
            producto_sku=codigo,
            producto_desc=clean_text(get_cell(row, idx_producto)),
            indicador=clean_text(get_cell(row, idx_indicador)),
            grupo=clean_text(get_cell(row, idx_grupo)),
            cantidad_pendiente=cantidad_pendiente,
            estatus_compras=clean_text(get_cell(row, idx_estatus)),
            disponibilidad_vl06o=None,
            cantidad_disponible=0.0,
            fecha_disponibilidad=None,
            dias_disponible=None,
            estado_crm=status_result.estado_crm,
            motivo_estado=status_result.motivo_estado,
        )
        db.add(record)
        rows_added += 1

    return rows_added


async def _replace_por_entregar_from_vl06o(db, ws, users_by_code):
    headers = [cell.value for cell in ws[1]]
    idx_factura = find_col_index(headers, [r"^factura$"])
    idx_codigo = find_col_index(headers, [r"^codigo$", r"c[oÃ³]digo"])
    idx_producto = find_col_index(headers, [r"producto"])
    idx_cantidad = find_col_index(headers, [r"cantidad.*entregar"])
    idx_vendedor = find_col_index(headers, [r"clave.*vendedor", r"vendedor"])
    idx_num_cliente = find_col_index(headers, [r"num.*cliente"])
    idx_cliente = find_col_index(headers, [r"nombre.*cliente"])
    idx_fecha = find_col_index(headers, [r"fecha.*disponibilidad"])
    idx_dias = find_col_index(headers, [r"dias.*disponible", r"d[iÃ­]as.*disponible"])

    await db.execute(delete(PorEntregar))

    rows_added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        factura = normalize_key(get_cell(row, idx_factura))
        codigo = normalize_key(get_cell(row, idx_codigo))
        cantidad_entregar = safe_float(get_cell(row, idx_cantidad))
        if not factura or not codigo or cantidad_entregar <= 0:
            continue

        vendedor_codigo = normalize_key(get_cell(row, idx_vendedor))
        vendedor = users_by_code.get(vendedor_codigo)
        vendedor_nombre = vendedor.nombre_completo if vendedor else vendedor_codigo
        dias_disponible = safe_int(get_cell(row, idx_dias))
        status_result = classify_por_entregar(dias_disponible)

        record = PorEntregar(
            factura=factura,
            producto_sku=codigo,
            producto_desc=clean_text(get_cell(row, idx_producto)),
            cantidad_entregar=cantidad_entregar,
            vendedor_codigo=vendedor_codigo,
            vendedor_nombre=vendedor_nombre,
            numero_cliente=normalize_key(get_cell(row, idx_num_cliente)),
            cliente_nombre=clean_text(get_cell(row, idx_cliente)),
            fecha_disponibilidad=parse_date(get_cell(row, idx_fecha)),
            dias_disponible=dias_disponible,
            estado_crm=status_result.estado_crm,
            motivo_estado=status_result.motivo_estado,
        )
        db.add(record)
        rows_added += 1

    return rows_added


def _availability_label(exact_matches, factura_matches):
    if exact_matches:
        return "Coincidencia exacta Factura + Codigo"
    if factura_matches:
        return "Factura relacionada en VL06O sin coincidencia de Codigo"
    return "Sin evidencia en VL06O"


def _first_non_empty(values):
    for value in values:
        if value is not None and value != "":
            return value
    return None
