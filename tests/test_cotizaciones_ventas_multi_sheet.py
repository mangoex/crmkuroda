import io
from datetime import date, time, datetime
from decimal import Decimal
from unittest.mock import AsyncMock, patch
import unittest
from uuid import uuid4
import openpyxl

from app.models.cotizacion import Cotizacion
from app.models.cotizacion_detalle import CotizacionItem
from app.models.promocion import Promocion
from app.models.usuario import Usuario
from app.api.v1.cotizaciones import (
    process_excel_background,
    generate_excel_template_bytes,
    is_multi_sheet_quote_format,
    serialize_cotizacion,
)
from app.services.commercial_analytics import get_promociones_intelligence_summary


def create_test_multisheet_excel_bytes() -> bytes:
    wb = openpyxl.Workbook()
    # Sheet 1: Ventas
    ws_ventas = wb.active
    ws_ventas.title = "Ventas"
    ws_ventas.append([
        "Fecha de Factura",
        "Numero del Cliente",
        "Plazo de Entrega",
        "Nombre del Cliente",
        "Folio Cotizacion",
        "Folio Factura",
        "Hora de Facturacion",
        "Margen",
        "Grupo de Vendedores",
        "Nombre del Vendedor",
        "Canal de Distribucion",
    ])
    ws_ventas.append([
        datetime(2026, 1, 15),
        "400191",
        "ENTREGA INMEDIATA",
        "PRODIVERSO CASA KURODA",
        "COT-1001",
        "FAC-9001",
        time(10, 30, 0),
        35.5,
        "C82",
        "Juan Perez",
        "01",
    ])
    ws_ventas.append([
        datetime(2026, 1, 16),
        "400200",
        "SOBREPEDIDO",
        "CONSTRUCTORA DEL NORTE",
        "COT-1002",
        "FAC-9002",
        time(14, 15, 0),
        28.0,
        "C94",
        "Maria Lopez",
        "02",
    ])

    # Sheet 2: Cotizaciones
    ws_cotizaciones = wb.create_sheet(title="Cotizaciones")
    ws_cotizaciones.append([
        "Fecha de Registro",
        "Organizacion de Ventas",
        "Numero de Cotizacion",
        "Indicador ABC+Frecuencia de Venta",
        "Codigo de Material",
        "Descripcion del Material",
        "Unidad de Medida",
        "Precio de Venta",
    ])
    # Items for COT-1001 (Partida 1: SKU Promo, Partida 2: Normal)
    ws_cotizaciones.append([
        datetime(2026, 1, 10),
        "MK01",
        "COT-1001",
        "C6",
        "SKU-PROMO-01",
        "VALVULA RETENCION 25MM",
        "PZA",
        150.00,
    ])
    ws_cotizaciones.append([
        datetime(2026, 1, 10),
        "MK01",
        "COT-1001",
        "A1",
        "SKU-NORMAL-01",
        "TUBO PVC HIDRAULICO 2 PULG",
        "TRAMO",
        250.00,
    ])
    # Items for COT-1002
    ws_cotizaciones.append([
        datetime(2026, 1, 12),
        "MK01",
        "COT-1002",
        "B2",
        "SKU-NORMAL-02",
        "PISO CERAMICO 60X60",
        "M2",
        300.00,
    ])
    # COT-1003 is a pending/lost quote without invoice in Ventas
    ws_cotizaciones.append([
        datetime(2026, 1, 14),
        "MK01",
        "COT-1003",
        "D6",
        "SKU-PROMO-02",
        "INSERTO DECORATIVO BEIGE",
        "PZA",
        80.00,
    ])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


class FakeScalarResult:
    def __init__(self, values):
        self.values = values

    def scalars(self):
        return self

    def all(self):
        return list(self.values)

    def first(self):
        return self.values[0] if self.values else None


class FakeAsyncSession:
    def __init__(self, users=None, promos=None, clients=None, quotes=None):
        self.users = users or []
        self.promos = promos or []
        self.clients = clients or []
        self.quotes = quotes or []
        self.added = []
        self.added_all_list = []
        self.deleted = []
        self.commits = 0
        self.flushes = 0

    async def execute(self, statement, params=None):
        if params and isinstance(params, list):
            self.added_all_list.extend([CotizacionItem(**p) if isinstance(p, dict) else p for p in params])
        str_stmt = str(statement).lower()
        if "from promociones" in str_stmt:
            return FakeScalarResult(self.promos)
        elif "from usuarios" in str_stmt:
            return FakeScalarResult(self.users)
        elif "from clientes" in str_stmt:
            return FakeScalarResult(self.clients)
        elif "from cotizaciones" in str_stmt:
            return FakeScalarResult(self.quotes)
        return FakeScalarResult([])

    def add(self, item):
        self.added.append(item)

    def add_all(self, items):
        self.added_all_list.extend(items)

    async def flush(self):
        self.flushes += 1

    async def commit(self):
        self.commits += 1

    async def rollback(self):
        pass

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        pass


class TestCotizacionesVentasMultiSheet(unittest.IsolatedAsyncioTestCase):
    def test_detect_multisheet_format(self):
        excel_bytes = create_test_multisheet_excel_bytes()
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True)
        self.assertTrue(is_multi_sheet_quote_format(wb))

    def test_generate_excel_template_bytes(self):
        template_bytes = generate_excel_template_bytes()
        self.assertGreater(len(template_bytes), 0)

        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
        self.assertIn("Ventas", wb.sheetnames)
        self.assertIn("Cotizaciones", wb.sheetnames)
        self.assertIn("Instrucciones", wb.sheetnames)

        ws_v = wb["Ventas"]
        ventas_headers = [c.value for c in ws_v[1]]
        self.assertIn("Fecha de la Factura", ventas_headers)
        self.assertIn("Folio Cotizacion", ventas_headers)
        self.assertIn("Margen", ventas_headers)
        self.assertIn("Numero de Vendedor", ventas_headers)
        self.assertIn("Importe con IVA", ventas_headers)
        self.assertIn("Codigo de Material", ventas_headers)

        ws_c = wb["Cotizaciones"]
        cot_headers = [c.value for c in ws_c[1]]
        self.assertIn("Numero de Cotizacion", cot_headers)
        self.assertIn("Codigo de Material", cot_headers)
        self.assertIn("Precio de Venta", cot_headers)

    async def test_process_excel_background_multisheet(self):
        user_id = uuid4()
        user = Usuario(id=user_id, nombre_completo="Juan Perez", codigo_vendedor="C82", rol="vendedor")
        promo = Promocion(
            id=1,
            centro="MK01",
            codigo_material="SKU-PROMO-01",
            descripcion_material="VALVULA RETENCION 25MM",
            precio_promocion=120.0,
        )

        fake_session = FakeAsyncSession(users=[user], promos=[promo], quotes=[])

        with patch("app.core.database.SessionLocal", return_value=fake_session), \
             patch("app.services.actualizaciones_datos.registrar_actualizacion_datos", new=AsyncMock()):
            excel_bytes = create_test_multisheet_excel_bytes()
            err = await process_excel_background(excel_bytes, user_id)
            self.assertIsNone(err)

            # Check quotes added
            added_quotes = [item for item in fake_session.added if isinstance(item, Cotizacion)]
            self.assertEqual(len(added_quotes), 3)

            quote_1001 = next(q for q in added_quotes if q.numero_cotizacion == "COT-1001")
            self.assertEqual(quote_1001.cliente_nombre, "PRODIVERSO CASA KURODA")
            self.assertEqual(quote_1001.numero_factura, "FAC-9001")
            self.assertEqual(quote_1001.total, Decimal("400.00"))
            self.assertEqual(float(quote_1001.margen), 35.5)
            self.assertEqual(quote_1001.grupo_vendedores, "C82")
            self.assertEqual(quote_1001.plazo_entrega, "ENTREGA INMEDIATA")

            # Check line items added
            added_items = [item for item in fake_session.added_all_list if isinstance(item, CotizacionItem)]
            self.assertEqual(len(added_items), 4)

            promo_item = next(i for i in added_items if i.codigo_material == "SKU-PROMO-01")
            self.assertTrue(promo_item.es_promocion)
            self.assertEqual(promo_item.indicador_abcf, "C6")
            self.assertEqual(promo_item.unidad_medida, "PZA")
            self.assertEqual(promo_item.precio_venta, Decimal("150.00"))

            normal_item = next(i for i in added_items if i.codigo_material == "SKU-NORMAL-01")
            self.assertFalse(normal_item.es_promocion)

    async def test_promociones_intelligence_summary(self):
        class FakeSummaryResult:
            def __init__(self, rows):
                self.rows = rows

            def all(self):
                return self.rows

        q1 = Cotizacion(id=uuid4(), numero_cotizacion="C-1", numero_factura="F-1", vendedor_nombre="Juan Perez", total=Decimal("500"), importe_facturado=Decimal("500"))
        i1 = CotizacionItem(id=uuid4(), cotizacion_id=q1.id, codigo_material="SKU-PROMO-1", descripcion="PROMO 1", es_promocion=True, precio_venta=Decimal("200"), importe_cotizado=Decimal("200"), importe_facturado=Decimal("200"))

        q2 = Cotizacion(id=uuid4(), numero_cotizacion="C-2", numero_factura=None, vendedor_nombre="Maria Lopez", total=Decimal("300"), importe_facturado=None)
        i2 = CotizacionItem(id=uuid4(), cotizacion_id=q2.id, codigo_material="SKU-PROMO-1", descripcion="PROMO 1", es_promocion=True, precio_venta=Decimal("150"), importe_cotizado=Decimal("150"), importe_facturado=Decimal("0"))

        fake_db = AsyncMock()
        fake_db.execute.return_value = FakeSummaryResult([(i1, q1), (i2, q2)])

        metrics = await get_promociones_intelligence_summary(fake_db)
        self.assertEqual(metrics["total_cotizaciones_con_promo"], 2)
        self.assertEqual(metrics["total_cotizaciones_promo_facturadas"], 1)
        self.assertEqual(metrics["total_partidas_promo"], 2)
        self.assertEqual(metrics["total_monto_cotizado_promo"], 350.0)
        self.assertEqual(metrics["total_monto_facturado_promo"], 200.0)
        self.assertEqual(metrics["tasa_efectividad_promo"], 50.0)
        self.assertEqual(len(metrics["top_skus_promo"]), 1)
        self.assertEqual(metrics["top_skus_promo"][0]["codigo_material"], "SKU-PROMO-1")

    def test_serialize_cotizacion_resumen_contains_all_multisheet_fields(self):
        quote = Cotizacion(
            id=uuid4(),
            cliente_nombre="Cliente Test",
            numero_cliente="12345",
            datos_contacto={"telefono": "6671234567"},
            total=Decimal("1500.00"),
            numero_cotizacion="COT-TEST",
            fecha_registro=date(2026, 1, 15),
            canal="01",
            numero_factura="FAC-123",
            fecha_factura=date(2026, 1, 16),
            hora_facturacion="10:30:00",
            margen=Decimal("35.500"),
            grupo_vendedores="C82",
            plazo_entrega="ENTREGA INMEDIATA",
            importe_facturado=Decimal("1500.00"),
            venta_perdida="No",
            comentarios="Comentario de prueba",
        )
        serialized = serialize_cotizacion(quote, vista="resumen")
        self.assertEqual(serialized["hora_facturacion"], "10:30:00")
        self.assertEqual(serialized["margen"], 35.5)
        self.assertEqual(serialized["grupo_vendedores"], "C82")
        self.assertEqual(serialized["plazo_entrega"], "ENTREGA INMEDIATA")
        self.assertEqual(serialized["numero_factura"], "FAC-123")
        self.assertEqual(serialized["importe_facturado"], 1500.00)
        self.assertEqual(serialized["materiales_cotizados"], None)

    async def test_multisheet_preserves_existing_client_and_advisor_for_uninvoiced_quote(self):
        user_id = uuid4()
        user = Usuario(id=user_id, nombre_completo="ARMANDO ARIAS", codigo_vendedor="O88", rol="vendedor")
        
        # Pre-existing quote in database with known client and seller
        existing_quote = Cotizacion(
            id=uuid4(),
            numero_cotizacion="COT-1003",
            cliente_nombre="LAMBERTO URIARTE",
            numero_cliente="400460",
            vendedor_id=user_id,
            vendedor_nombre="ARMANDO ARIAS",
            datos_contacto={"telefono": "6671234567", "email": "lamberto@example.com"},
            canal="01",
            total=Decimal("80.00"),
        )

        fake_session = FakeAsyncSession(users=[user], promos=[], quotes=[existing_quote])

        with patch("app.core.database.SessionLocal", return_value=fake_session), \
             patch("app.services.actualizaciones_datos.registrar_actualizacion_datos", new=AsyncMock()):
            # Excel has COT-1003 only in Cotizaciones sheet (not in Ventas)
            excel_bytes = create_test_multisheet_excel_bytes()
            err = await process_excel_background(excel_bytes, user_id)
            self.assertIsNone(err)

            # COT-1003 should preserve its client and seller!
            self.assertEqual(existing_quote.cliente_nombre, "LAMBERTO URIARTE")
            self.assertEqual(existing_quote.numero_cliente, "400460")
            self.assertEqual(existing_quote.vendedor_nombre, "ARMANDO ARIAS")
            self.assertEqual(existing_quote.vendedor_id, user_id)
            self.assertEqual(existing_quote.materiales_cotizados, "SKU-PROMO-02")
            self.assertEqual(len(existing_quote.items), 1)

    async def test_multisheet_with_optional_cotizaciones_columns(self):
        user_id = uuid4()
        user = Usuario(id=user_id, nombre_completo="LORENA PERAZA", codigo_vendedor="C01", rol="vendedor")

        wb = openpyxl.Workbook()
        ws_ventas = wb.active
        ws_ventas.title = "Ventas"
        ws_ventas.append(["Fecha de Factura", "Numero del Cliente", "Plazo de Entrega", "Nombre del Cliente", "Folio Cotizacion", "Folio Factura", "Hora de Facturacion", "Margen", "Grupo de Vendedores", "Nombre del Vendedor", "Canal de Distribucion"])
        # Ventas is empty (no invoices yet)

        ws_c = wb.create_sheet(title="Cotizaciones")
        ws_c.append([
            "Fecha de Registro", "Organizacion de Ventas", "Numero de Cotizacion", "Indicador ABC+Frecuencia de Venta",
            "Codigo de Material", "Descripcion del Material", "Unidad de Medida", "Precio de Venta",
            "Nombre del Cliente", "Numero del Cliente", "Nombre del Vendedor", "Vendedor", "Canal"
        ])
        ws_c.append([
            datetime(2026, 8, 20), "MK01", "COT-9999", "A1",
            "TUBO-PVC-4", "TUBO PVC 4 PULG", "TRAMO", 350.00,
            "CONSTRUCTORA EJEMPLO SA", "554433", "LORENA PERAZA", "C01", "02"
        ])

        buf = io.BytesIO()
        wb.save(buf)
        excel_bytes = buf.getvalue()

        fake_session = FakeAsyncSession(users=[user], promos=[], quotes=[])

        with patch("app.core.database.SessionLocal", return_value=fake_session), \
             patch("app.services.actualizaciones_datos.registrar_actualizacion_datos", new=AsyncMock()):
            err = await process_excel_background(excel_bytes, user_id)
            self.assertIsNone(err)

            added_quotes = [item for item in fake_session.added if isinstance(item, Cotizacion)]
            self.assertEqual(len(added_quotes), 1)
            q = added_quotes[0]
            self.assertEqual(q.numero_cotizacion, "COT-9999")
            self.assertEqual(q.cliente_nombre, "CONSTRUCTORA EJEMPLO SA")
            self.assertEqual(q.numero_cliente, "554433")
            self.assertEqual(q.vendedor_nombre, "LORENA PERAZA")
            self.assertEqual(q.vendedor_id, user_id)
            self.assertEqual(q.canal, "02")
            self.assertEqual(q.materiales_cotizados, "TUBO-PVC-4")
            self.assertEqual(len(q.items), 1)

    async def test_multisheet_with_arbitrary_sheet_names_cotizaciones_and_detalle(self):
        user_id = uuid4()
        user = Usuario(id=user_id, nombre_completo="CARLOS VALENZUELA", codigo_vendedor="C50", rol="vendedor")

        wb = openpyxl.Workbook()
        # Hoja 1: Cotizaciones (cabecera con clientes y vendedores)
        ws_cot = wb.active
        ws_cot.title = "Cotizaciones"
        ws_cot.append(["Fecha de Registro", "Numero de Cotizacion", "Nombre del Cliente", "Numero del Cliente", "Nombre del Vendedor", "Vendedor", "Canal"])
        ws_cot.append([datetime(2026, 8, 25), "416787053", "DISTRIBUIDORA DEL PACIFICO", "400888", "CARLOS VALENZUELA", "C50", "01"])

        # Hoja 2: Detalle de Cotizacion (renglones con productos)
        ws_det = wb.create_sheet(title="Detalle de Cotizacion")
        ws_det.append(["Numero de Cotizacion", "Codigo de Material", "Descripcion del Material", "Unidad de Medida", "Precio de Venta", "Indicador ABC+Frecuencia de Venta"])
        ws_det.append(["416787053", "ST101", "TUBO PVC SANIT NORMA 101MM X 6MT", "TM8", 307.14, "A6"])
        ws_det.append(["416787053", "ST152", "TUBO PVC SANIT NORMA 152MM X 6MT", "TM8", 877.20, "A6"])

        buf = io.BytesIO()
        wb.save(buf)
        excel_bytes = buf.getvalue()

        fake_session = FakeAsyncSession(users=[user], promos=[], quotes=[])

        with patch("app.core.database.SessionLocal", return_value=fake_session), \
             patch("app.services.actualizaciones_datos.registrar_actualizacion_datos", new=AsyncMock()):
            err = await process_excel_background(excel_bytes, user_id)
            self.assertIsNone(err)

            added_quotes = [item for item in fake_session.added if isinstance(item, Cotizacion)]
            self.assertEqual(len(added_quotes), 1)
            q = added_quotes[0]
            self.assertEqual(q.numero_cotizacion, "416787053")
            self.assertEqual(q.cliente_nombre, "DISTRIBUIDORA DEL PACIFICO")
            self.assertEqual(q.numero_cliente, "400888")
            self.assertEqual(q.vendedor_nombre, "CARLOS VALENZUELA")
            self.assertEqual(q.vendedor_id, user_id)
            self.assertEqual(q.total, Decimal("1184.34"))
            self.assertEqual(q.materiales_cotizados, "ST101, ST152")
            self.assertEqual(len(q.items), 2)

    async def test_process_excel_background_definitivo_16_cols(self):
        user_id = uuid4()
        user = Usuario(id=user_id, nombre_completo="Aaron Emigdio Lechuga", codigo_vendedor="C82", rol="vendedor")
        promo = Promocion(
            id=1,
            centro="MK01",
            codigo_material="CFIP137",
            descripcion_material="LLAVE ANGULAR 2 SALIDAS 1/2X1/2 COFLEX",
            precio_promocion=140.00,
        )

        wb = openpyxl.Workbook()
        # Sheet 1: Ventas (16 columnas definitivas con partidas a nivel renglón)
        ws_v = wb.active
        ws_v.title = "Ventas"
        ws_v.append([
            "Fecha de la Factura", "Numero del Cliente", "Plazo de Entrega", "Nombre del Cliente",
            "Folio Cotizacion", "Folio de la Factura", "Hora de la Factura", "Codigo de Material",
            "Descripcion del Material", "Cantidad Facturada UMB", "Importe con IVA", "Margen",
            "Numero de Vendedor", "Nombre del Vendedor", "Indicador ABC+Frecuencia de Venta", "Canal de Distribucion"
        ])
        # Cotización 416662481: Facturada con 2 renglones
        ws_v.append([
            datetime(2026, 1, 2), "400191", "ENTREGA INMEDIATA", "ADAM",
            "416662481", "1325607092", time(9, 52, 16), "CFIP137",
            "LLAVE ANGULAR 2 SALIDAS 1/2X1/2 COFLEX", 1, 158.02, 35.81,
            "C82", "Aaron Emigdio Lechuga", "B3", "01"
        ])
        ws_v.append([
            datetime(2026, 1, 2), "400191", "ENTREGA INMEDIATA", "ADAM",
            "416662481", "1325607092", time(9, 52, 16), "4119",
            "LLAVE NARIZ P/MANGUERA 13MM DICA", 2, 178.00, 29.75,
            "C82", "Aaron Emigdio Lechuga", "A2", "01"
        ])
        # Cotización 416662488: Facturación parcial (cotizó 2 productos, solo compró 1)
        ws_v.append([
            datetime(2026, 1, 3), "400200", "SOBREPEDIDO", "CONSTRUCTORA DEL NORTE",
            "416662488", "1325607095", time(11, 20, 0), "TUBOPVC2",
            "TUBO PVC HIDRAULICO 2 PULG", 1, 200.00, 25.00,
            "C82", "Aaron Emigdio Lechuga", "A1", "01"
        ])

        # Sheet 2: Cotizaciones (8 columnas con partidas cotizadas)
        ws_c = wb.create_sheet(title="Cotizaciones")
        ws_c.append([
            "Fecha de Registro", "Organizacion de Ventas", "Numero de Cotizacion", "Indicador ABC+Frecuencia de Venta",
            "Codigo de Material", "Descripcion del Material", "Unidad de Medida", "Precio de Venta"
        ])
        # Cotización 416662481: Cotizó CFIP137 ($150) y 4119 ($170) -> Total Cotizado = $320
        ws_c.append([
            datetime(2026, 1, 2), "MK01", "416662481", "B3",
            "CFIP137", "LLAVE ANGULAR 2 SALIDAS 1/2X1/2 COFLEX", "PZA", 150.00
        ])
        ws_c.append([
            datetime(2026, 1, 2), "MK01", "416662481", "A2",
            "4119", "LLAVE NARIZ P/MANGUERA 13MM DICA", "PZA", 170.00
        ])
        # Cotización 416662488: Cotizó TUBOPVC2 ($200) y CODO90 ($100) -> Total Cotizado = $300
        ws_c.append([
            datetime(2026, 1, 3), "MK01", "416662488", "A1",
            "TUBOPVC2", "TUBO PVC HIDRAULICO 2 PULG", "TRAMO", 200.00
        ])
        ws_c.append([
            datetime(2026, 1, 3), "MK01", "416662488", "A3",
            "CODO90", "CODO 90 SANITARIO 2 PULG", "PZA", 100.00
        ])
        # Cotización 416662499: Cotización viva (no facturada)
        ws_c.append([
            datetime(2026, 1, 4), "MK01", "416662499", "C1",
            "PEGAMENTOPVC", "PEGAMENTO PVC 250ML", "BOTE", 85.00
        ])

        buf = io.BytesIO()
        wb.save(buf)
        excel_bytes = buf.getvalue()

        fake_session = FakeAsyncSession(users=[user], promos=[promo], quotes=[])

        with patch("app.core.database.SessionLocal", return_value=fake_session), \
             patch("app.services.actualizaciones_datos.registrar_actualizacion_datos", new=AsyncMock()):
            err = await process_excel_background(excel_bytes, user_id)
            self.assertIsNone(err)

            added_quotes = {q.numero_cotizacion: q for q in fake_session.added if isinstance(q, Cotizacion)}
            self.assertEqual(len(added_quotes), 3)

            # Cotización 416662481 (Totalmente Vendida)
            q1 = added_quotes["416662481"]
            self.assertEqual(q1.cliente_nombre, "ADAM")
            self.assertEqual(q1.numero_cliente, "400191")
            self.assertEqual(q1.numero_factura, "1325607092")
            self.assertEqual(q1.total, Decimal("320.00"))
            self.assertEqual(q1.importe_facturado, Decimal("336.02"))  # 158.02 + 178.00
            self.assertEqual(q1.materiales_cotizados, "CFIP137, 4119")
            self.assertEqual(q1.materiales_facturados, "CFIP137, 4119")
            self.assertEqual(q1.porcentaje_materiales, Decimal("100.00"))
            self.assertEqual(q1.porcentaje_importe, Decimal("105.01"))
            self.assertEqual(len(q1.items), 2)
            # Verificar items en JSON
            items_by_code = {it["codigo_material"]: it for it in q1.items}
            self.assertEqual(items_by_code["CFIP137"]["cantidad_facturada"], 1.0)
            self.assertEqual(items_by_code["CFIP137"]["importe_facturado"], 158.02)
            self.assertTrue(items_by_code["CFIP137"]["es_promocion"])
            self.assertEqual(items_by_code["4119"]["cantidad_facturada"], 2.0)
            self.assertEqual(items_by_code["4119"]["importe_facturado"], 178.00)

            # Cotización 416662488 (Parcialmente Vendida)
            q2 = added_quotes["416662488"]
            self.assertEqual(q2.total, Decimal("300.00"))
            self.assertEqual(q2.importe_facturado, Decimal("200.00"))
            self.assertEqual(q2.materiales_cotizados, "TUBOPVC2, CODO90")
            self.assertEqual(q2.materiales_facturados, "TUBOPVC2")
            self.assertEqual(q2.porcentaje_materiales, Decimal("50.00"))
            self.assertEqual(q2.porcentaje_importe, Decimal("66.67"))
            q2_items = {it["codigo_material"]: it for it in q2.items}
            self.assertEqual(q2_items["TUBOPVC2"]["cantidad_facturada"], 1.0)
            self.assertEqual(q2_items["CODO90"]["cantidad_facturada"], 0.0)
            self.assertEqual(q2_items["CODO90"]["importe_facturado"], 0.0)

            # Cotización 416662499 (Viva / Sin factura)
            q3 = added_quotes["416662499"]
            self.assertIsNone(q3.numero_factura)
            self.assertIsNone(q3.importe_facturado)
            self.assertEqual(q3.total, Decimal("85.00"))
            self.assertEqual(q3.materiales_cotizados, "PEGAMENTOPVC")
            self.assertIsNone(q3.materiales_facturados)
            self.assertIsNone(q3.porcentaje_materiales)
            self.assertIsNone(q3.porcentaje_importe)

    def test_serialize_cotizacion_resilience_to_missing_attributes(self):
        # Simular una instancia con campos parciales (como ocurre con load_only)
        q = Cotizacion(
            id=uuid4(),
            numero_cotizacion="COT-TEST-RESILIENCE",
            cliente_nombre="CLIENTE PRUEBA",
            total=Decimal("500.00"),
        )
        # Serializar en vista resumen y vista completa
        data_resumen = serialize_cotizacion(q, vista="resumen")
        self.assertEqual(data_resumen["numero_cotizacion"], "COT-TEST-RESILIENCE")
        self.assertEqual(data_resumen["cliente_nombre"], "CLIENTE PRUEBA")
        self.assertIsNone(data_resumen["materiales_facturados"])
        self.assertIsNone(data_resumen["porcentaje_materiales"])
        self.assertNotIn("items", data_resumen)

        data_completa = serialize_cotizacion(q, vista="completa")
        self.assertEqual(data_completa["numero_cotizacion"], "COT-TEST-RESILIENCE")
        self.assertIn("items", data_completa)

    def test_serialize_cotizacion_resolves_unknown_client_from_client_number(self):
        # Cotización con 'Cliente Desconocido' pero con numero_cliente
        q = Cotizacion(
            id=uuid4(),
            numero_cotizacion="COT-416788516",
            cliente_nombre="Cliente Desconocido",
            numero_cliente="400460",
            total=Decimal("738.90"),
        )
        serialized = serialize_cotizacion(q, vista="resumen")
        self.assertEqual(serialized["cliente_nombre"], "Cliente #400460")

    def test_serialize_cotizacion_resolves_unknown_client_from_contact_name(self):
        # Cotización con 'Cliente Desconocido' pero con nombre_contacto en datos_contacto
        q = Cotizacion(
            id=uuid4(),
            numero_cotizacion="COT-416788515",
            cliente_nombre="Cliente Desconocido",
            datos_contacto={"nombre_contacto": "ING. ROBERTO SALAZAR", "telefono": "6671234567"},
            total=Decimal("203.79"),
        )
        serialized = serialize_cotizacion(q, vista="resumen")
        self.assertEqual(serialized["cliente_nombre"], "ING. ROBERTO SALAZAR")

    def test_serialize_cotizacion_resolves_unknown_client_from_enrichment(self):
        # Cotización con 'Cliente Desconocido' enriquecida desde catálogo Cliente
        q = Cotizacion(
            id=uuid4(),
            numero_cotizacion="COT-416788513",
            cliente_nombre="Cliente Desconocido",
            total=Decimal("1056.55"),
        )
        enrichment = {"cliente_nombre": "CONSTRUCTORA PACIFICO SUR"}
        serialized = serialize_cotizacion(q, enrichment=enrichment, vista="resumen")
        self.assertEqual(serialized["cliente_nombre"], "CONSTRUCTORA PACIFICO SUR")






