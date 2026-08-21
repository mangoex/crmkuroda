import io
import unittest
from datetime import datetime, timezone
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch
from uuid import uuid4

import openpyxl
from fastapi import HTTPException, UploadFile

from app.api.v1.cotizaciones import (
    _apply_imported_quote_values,
    _get_authorized_quote,
    create_quote_comment,
    process_excel_background,
    update_quote_comment,
    upload_quote_material_detail,
)
from app.api.v1.commercial_analytics import sales_by_channel
from app.schemas.commercial import ComentarioCreate, ComentarioUpdate


class FakeScalarResult:
    def __init__(self, values):
        self.values = values

    def scalars(self):
        return self

    def all(self):
        return list(self.values)

    def first(self):
        return self.values[0] if self.values else None


class FakeDatabase:
    def __init__(self, execute_results):
        self.execute_results = list(execute_results)
        self.added = []
        self.added_many = []
        self.commits = 0
        self.rollbacks = 0
        self.execute_calls = 0

    async def execute(self, _statement):
        self.execute_calls += 1
        if self.execute_results:
            return self.execute_results.pop(0)
        return FakeScalarResult([])

    def add(self, item):
        self.added.append(item)

    def add_all(self, items):
        self.added_many.extend(items)

    async def commit(self):
        self.commits += 1

    async def flush(self):
        pass

    async def rollback(self):
        self.rollbacks += 1

    async def refresh(self, item):
        if item.id is None:
            item.id = uuid4()
        if getattr(item, "creado_en", None) is None:
            item.creado_en = datetime.now(timezone.utc).replace(tzinfo=None)


def build_detail_workbook(rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Numero de Cotizacion",
            "Codigo Material",
            "Descripcion",
            "Familia",
            "Grupo de Materiales",
            "Cantidad Cotizada",
            "Importe Cotizado",
            "Cantidad Facturada",
            "Importe Facturado",
        ]
    )
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_summary_workbook(
    *,
    tipo_entrega="PIDE Y RECOGE",
    reorder_columns=False,
    omit_header=None,
):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    pairs = [
        ("Fecha de Registro", datetime(2026, 8, 10)),
        ("Organizacion de Ventas", "MK01"),
        ("Numero de Cotizacion", "416662973"),
        ("Canal", "01"),
        ("Vendedor", "C01"),
        ("Nombre del Vendedor", "LORENA PERAZA"),
        ("Numero del Cliente", "400191"),
        ("Nombre del Cliente", "JESUS JAIME HERNANDEZ CAM"),
        ("Numero de Telefono", "1"),
        ("Numero de Celular", None),
        ("Direccion Correo Electronico", "cliente@example.com"),
        ("Numero de Factura", "1325607139"),
        ("Fecha de Factura", datetime(2026, 8, 10)),
        ("Importe Cotizado c/IVA", 751.44),
        ("Importe Facturado c/IVA", 751.44),
        ("Porcentaje de Importe", 100),
        ("Materiales Cotizados", 2),
        ("Materiales Facturados", 2),
        ("Porcentaje de Materiales", 100),
    ]
    if tipo_entrega is not None:
        pairs.append(("Tipo de Entrega", tipo_entrega))
    if omit_header is not None:
        pairs = [pair for pair in pairs if pair[0] != omit_header]
    if reorder_columns:
        pairs = [pairs[-1], *pairs[:-1]]
    sheet.append([header for header, _value in pairs])
    sheet.append([value for _header, value in pairs])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class FakeSessionLocal:
    def __init__(self, database):
        self.database = database

    async def __aenter__(self):
        return self.database

    async def __aexit__(self, exc_type, exc, traceback):
        return False


class CommercialApiWorkflowTest(unittest.IsolatedAsyncioTestCase):
    async def test_summary_import_uses_tipo_entrega_as_channel(self):
        database = FakeDatabase(
            [
                FakeScalarResult([]),
                FakeScalarResult([]),
            ]
        )
        register_update = AsyncMock()

        with (
            patch(
                "app.core.database.SessionLocal",
                return_value=FakeSessionLocal(database),
            ),
            patch(
                "app.services.actualizaciones_datos.registrar_actualizacion_datos",
                new=register_update,
            ),
        ):
            error = await process_excel_background(
                build_summary_workbook(
                    tipo_entrega="PIDE Y RECOGE",
                    reorder_columns=True,
                ),
                uuid4(),
            )

        self.assertIsNone(error)
        self.assertEqual(database.added_many[0].canal, "PIDE Y RECOGE")
        self.assertEqual(database.added_many[0].numero_cotizacion, "416662973")
        self.assertEqual(database.commits, 1)
        register_update.assert_awaited_once()

    async def test_summary_import_accepts_legacy_canal_column(self):
        database = FakeDatabase(
            [
                FakeScalarResult([]),
                FakeScalarResult([]),
            ]
        )

        with (
            patch(
                "app.core.database.SessionLocal",
                return_value=FakeSessionLocal(database),
            ),
            patch(
                "app.services.actualizaciones_datos.registrar_actualizacion_datos",
                new=AsyncMock(),
            ),
        ):
            error = await process_excel_background(
                build_summary_workbook(tipo_entrega=None),
                uuid4(),
            )

        self.assertIsNone(error)
        self.assertEqual(database.added_many[0].canal, "01")

    async def test_summary_import_rejects_incomplete_layout_before_queries(self):
        database = FakeDatabase([])

        with patch(
            "app.core.database.SessionLocal",
            return_value=FakeSessionLocal(database),
        ):
            error = await process_excel_background(
                build_summary_workbook(omit_header="Nombre del Cliente"),
                uuid4(),
            )

        self.assertIn("Faltan columnas requeridas: NOMBRE DEL CLIENTE", error)

    async def test_summary_import_multi_sheet_selects_quote_sheet(self):
        wb = openpyxl.Workbook()
        ws_extra = wb.active
        ws_extra.title = "Hoja1"
        ws_extra.append(["Folio Factura", "Plazo de Entrega"])
        ws_extra.append(["FAC-01", "30 dias"])

        ws_quotes = wb.create_sheet(title="Sheet1")
        ws_quotes.append([
            "Fecha de Registro", "Organizacion de Ventas", "Numero de Cotizacion", "Canal",
            "Vendedor", "Nombre del Vendedor", "Numero del Cliente", "Nombre del Cliente",
            "Numero de Telefono", "Numero de Celular", "Direccion Correo Electronico",
            "Numero de Factura", "Fecha de Factura", "Importe Cotizado c/IVA",
            "Importe Facturado c/IVA", "Porcentaje de Importe", "Materiales Cotizados",
            "Materiales Facturados", "Porcentaje de Materiales", "Tipo de Entrega"
        ])
        ws_quotes.append([
            "2026-01-02", "MK01", "123456", "01", "V01", "Juan Perez", "C01", "Cliente Test",
            "5551234", "5555678", "test@example.com", None, None, "1000", "0", "0", "1", "0", "0", "PIDE Y RECOGE"
        ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        database = FakeDatabase([FakeScalarResult([]), FakeScalarResult([])])

        with (
            patch("app.core.database.SessionLocal", return_value=FakeSessionLocal(database)),
            patch("app.services.actualizaciones_datos.registrar_actualizacion_datos", new=AsyncMock()),
        ):
            error = await process_excel_background(buffer.getvalue(), uuid4())

        self.assertIsNone(error)
        self.assertEqual(len(database.added_many), 1)
        self.assertEqual(database.added_many[0].numero_cotizacion, "123456")
        self.assertEqual(database.commits, 1)
        self.assertEqual(database.rollbacks, 0)

    async def test_summary_reconciliation_preserves_followup_fields(self):
        quote_id = uuid4()
        quote = SimpleNamespace(
            id=quote_id,
            numero_cotizacion="COT-1",
            total=100,
            comentarios="motivo histórico",
            comentarios_seguimiento=["seguimiento"],
        )

        result = _apply_imported_quote_values(
            quote,
            {
                "numero_cotizacion": "COT-1",
                "total": 250,
                "cliente_nombre": "Cliente actualizado",
            },
        )

        self.assertIs(result, quote)
        self.assertEqual(result.id, quote_id)
        self.assertEqual(result.total, 250)
        self.assertEqual(result.comentarios, "motivo histórico")
        self.assertEqual(result.comentarios_seguimiento, ["seguimiento"])

    async def test_support_role_cannot_access_commercial_analytics(self):
        user = SimpleNamespace(id=uuid4(), rol="soporte")
        database = FakeDatabase([])

        with self.assertRaises(HTTPException) as raised:
            await sales_by_channel(
                fecha_inicio=None,
                fecha_fin=None,
                vendedor_id=None,
                sin_vincular=False,
                db=database,
                current_user=user,
            )

        self.assertEqual(raised.exception.status_code, 403)
        self.assertEqual(database.execute_calls, 0)

    async def test_support_role_cannot_add_followup_comments(self):
        user = SimpleNamespace(id=uuid4(), rol="soporte")
        database = FakeDatabase([])

        with self.assertRaises(HTTPException) as raised:
            await create_quote_comment(
                uuid4(),
                ComentarioCreate(comentario="No permitido"),
                database,
                user,
            )

        self.assertEqual(raised.exception.status_code, 403)
        self.assertEqual(database.execute_calls, 0)

    async def test_seller_can_open_historical_quote_linked_by_name(self):
        user_id = uuid4()
        quote = SimpleNamespace(
            id=uuid4(),
            vendedor_id=None,
            vendedor_nombre="ANA ASESORA",
        )
        user = SimpleNamespace(
            id=user_id,
            rol="vendedor",
            codigo_vendedor="A01",
            nombre_completo="Ana Asesora",
            email="ana@example.com",
        )
        database = FakeDatabase(
            [
                FakeScalarResult([quote]),
                FakeScalarResult([]),
                FakeScalarResult([user]),
            ]
        )

        result = await _get_authorized_quote(database, user, quote.id)

        self.assertIs(result, quote)

    async def test_name_fallback_cannot_override_a_different_linked_seller(self):
        user_id = uuid4()
        quote = SimpleNamespace(
            id=uuid4(),
            vendedor_id=uuid4(),
            vendedor_nombre="Ana Asesora",
        )
        user = SimpleNamespace(
            id=user_id,
            rol="vendedor",
            codigo_vendedor="A01",
            nombre_completo="Ana Asesora",
            email="ana@example.com",
        )
        database = FakeDatabase(
            [
                FakeScalarResult([quote]),
                FakeScalarResult([]),
            ]
        )

        with self.assertRaises(HTTPException) as raised:
            await _get_authorized_quote(database, user, quote.id)

        self.assertEqual(raised.exception.status_code, 403)
        self.assertEqual(database.execute_calls, 2)

    async def test_material_upload_replaces_detail_only_after_valid_rows(self):
        quote_id = uuid4()
        quote = SimpleNamespace(id=quote_id, numero_cotizacion="COT-1")
        database = FakeDatabase([FakeScalarResult([quote]), FakeScalarResult([])])
        upload = UploadFile(
            filename="detalle.xlsx",
            file=build_detail_workbook(
                [["COT-1", "SKU-1", "Piso", "Acabados", "Pisos", 2, 200, 1, 100]]
            ),
        )
        user = SimpleNamespace(id=uuid4(), rol="gerente")

        result = await upload_quote_material_detail(upload, database, user)

        self.assertEqual(result["aceptadas"], 1)
        self.assertEqual(result["rechazadas"], 0)
        self.assertEqual(database.added_many[0].cotizacion_id, quote_id)
        self.assertEqual(database.added_many[0].codigo_material, "SKU-1")
        self.assertEqual(database.commits, 1)
        self.assertEqual(database.execute_calls, 2)

    async def test_invalid_material_upload_does_not_delete_existing_detail(self):
        quote = SimpleNamespace(id=uuid4(), numero_cotizacion="COT-1")
        database = FakeDatabase([FakeScalarResult([quote])])
        upload = UploadFile(
            filename="detalle.xlsx",
            file=build_detail_workbook(
                [["NO-EXISTE", "SKU-1", "Piso", "Acabados", "Pisos", 2, 200, 1, 100]]
            ),
        )
        user = SimpleNamespace(id=uuid4(), rol="gerente")

        with self.assertRaises(HTTPException) as raised:
            await upload_quote_material_detail(upload, database, user)

        self.assertEqual(raised.exception.status_code, 400)
        self.assertEqual(database.execute_calls, 1)
        self.assertEqual(database.commits, 0)
        self.assertEqual(database.added_many, [])

    async def test_followup_comment_is_independent_from_legacy_comments(self):
        quote = SimpleNamespace(
            id=uuid4(),
            vendedor_id=uuid4(),
            vendedor_nombre="Asesor",
            comentarios='{"lost_reason":{"reasons":["precio"]}}',
        )
        database = FakeDatabase([FakeScalarResult([quote])])
        user = SimpleNamespace(
            id=uuid4(),
            rol="gerente",
            nombre_completo="Gerente",
            email="gerente@example.com",
        )

        result = await create_quote_comment(
            quote.id,
            ComentarioCreate(comentario="  Llamar el lunes  "),
            database,
            user,
        )

        self.assertEqual(result["data"]["comentario"], "Llamar el lunes")
        self.assertEqual(
            quote.comentarios,
            '{"lost_reason":{"reasons":["precio"]}}',
        )
        self.assertEqual(database.commits, 1)

    async def test_comment_author_can_edit_without_changing_legacy_comment(self):
        user_id = uuid4()
        quote = SimpleNamespace(
            id=uuid4(),
            vendedor_id=user_id,
            vendedor_nombre="Ana",
            comentarios="motivo histórico",
        )
        comment = SimpleNamespace(
            id=uuid4(),
            cotizacion_id=quote.id,
            autor_id=user_id,
            comentario="Llamar",
            creado_en=datetime.now(timezone.utc).replace(tzinfo=None),
            editado_en=None,
        )
        user = SimpleNamespace(
            id=user_id,
            rol="vendedor",
            codigo_vendedor="A01",
            nombre_completo="Ana",
            email="ana@example.com",
        )
        database = FakeDatabase(
            [
                FakeScalarResult([quote]),
                FakeScalarResult([]),
                FakeScalarResult([comment]),
                FakeScalarResult([user]),
            ]
        )

        result = await update_quote_comment(
            quote.id,
            comment.id,
            ComentarioUpdate(comentario="  Visita confirmada  "),
            database,
            user,
        )

        self.assertEqual(result["data"]["comentario"], "Visita confirmada")
        self.assertEqual(quote.comentarios, "motivo histórico")
        self.assertIsNotNone(comment.editado_en)
        self.assertEqual(database.commits, 1)

    async def test_seller_cannot_edit_another_authors_comment(self):
        user_id = uuid4()
        quote = SimpleNamespace(id=uuid4(), vendedor_id=user_id, vendedor_nombre="Ana")
        comment = SimpleNamespace(
            id=uuid4(),
            cotizacion_id=quote.id,
            autor_id=uuid4(),
            comentario="No editable",
        )
        user = SimpleNamespace(
            id=user_id,
            rol="vendedor",
            codigo_vendedor="A01",
            nombre_completo="Ana",
            email="ana@example.com",
        )
        database = FakeDatabase(
            [
                FakeScalarResult([quote]),
                FakeScalarResult([]),
                FakeScalarResult([comment]),
            ]
        )

        with self.assertRaises(HTTPException) as raised:
            await update_quote_comment(
                quote.id,
                comment.id,
                ComentarioUpdate(comentario="Intento"),
                database,
                user,
            )

        self.assertEqual(raised.exception.status_code, 403)
        self.assertEqual(database.commits, 0)

    async def test_upload_promociones_multi_sheet_selects_promo_sheet(self):
        from app.api.v1.promociones import upload_promociones
        wb = openpyxl.Workbook()
        
        # Sheet 1: Inventario (8 cols)
        ws_inv = wb.active
        ws_inv.title = "Inventario"
        ws_inv.append(["Centro", "Codigo Material", "Nombre del Proveedor", "Descripcion Material", "Cantidad Propia", "Entregas Pendientes", "Existencia en Consignacion", "Inventario disponible"])
        ws_inv.append(["MK01", "FGA70719", "FUTURA", "CODO 90", "100", "0", "0", "100"])
        
        # Sheet 2: Promociones (16 cols)
        ws_promo = wb.create_sheet(title="Promociones")
        ws_promo.append([
            "Centro", "Descrip Gpo Materiales", "Indicador ABC+Frecuencia de Venta", "Codigo Material",
            "Descripcion del Material", "Unidad de Medida Base", "Costo Promedio Unitario",
            "Costo Promedio Unitario Moneda de Venta", "Costo Estandar", "Precio Efectivo Promocion",
            "Moneda", "Valido hasta Promocion", "Costo Estandar Promocion", "Margen Promocion",
            "Proveedor", "Inventario disponible"
        ])
        ws_promo.append([
            "MK01", "PLOMERIA", "A", "FGA70719", "CODO 90", "PZA", "10.0", "10.0", "12.0", "15.0", "MXN",
            "2026-12-31", "12.0", "20.0", "FUTURA", "100"
        ])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        upload_file = UploadFile(filename="promociones_multi.xlsx", file=buffer)
        user = SimpleNamespace(id=uuid4(), rol="admin")
        database = FakeDatabase([])

        with patch("app.api.v1.promociones.registrar_actualizacion_datos", new_callable=AsyncMock):
            result = await upload_promociones(upload_file, database, user)

        self.assertEqual(result["status"], "success")
        self.assertIn("1", result["message"])
        self.assertEqual(database.commits, 1)

    async def test_upload_promociones_csv(self):
        from app.api.v1.promociones import upload_promociones
        csv_content = (
            "Centro,Descrip Gpo Materiales,Indicador ABC,Codigo Material,Descripcion Material,Unidad Medida,"
            "Costo Promedio,Costo Promedio Moneda,Costo Estandar,Precio Promocion,Moneda,Valido Hasta,"
            "Costo Estandar Promocion,Margen Promocion,Proveedor,Inventario Disponible\n"
            "MK01,PLOMERIA,A,SKU123,CODO CPVC,PZA,10,10,12,15.50,MXN,2026-12-31,12,23.5,FUTURA,50\n"
        ).encode("utf-8")
        
        upload_file = UploadFile(filename="promociones.csv", file=io.BytesIO(csv_content))
        user = SimpleNamespace(id=uuid4(), rol="admin")
        database = FakeDatabase([])

        with patch("app.api.v1.promociones.registrar_actualizacion_datos", new_callable=AsyncMock):
            result = await upload_promociones(upload_file, database, user)

        self.assertEqual(result["status"], "success")
        self.assertIn("1", result["message"])
        self.assertEqual(database.commits, 1)


if __name__ == "__main__":
    unittest.main()

